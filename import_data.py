# -*- coding: utf-8 -*-
"""
import_data.py
קורא את קובץ סידור_עבודה_חדש.xlsx (עם לוח הרמזור)
ומכניס את כל הנתונים + הלוח לבסיס הנתונים
"""
import openpyxl
import sqlite3
import os
import sys

DB_PATH = os.environ.get("DB_PATH", "sales.db")

COLOR_TO_TL = {
    'FFFF0000': 'אדום',
    'FFFFC000': 'כתום',
    'FF92D050': 'ירוק',
}

# שם ביקור מלא → אות עברית (כמו שנשמר בDB ישן)
VISIT_DAY_MAP = {
    'ראשון':  'א',
    'שני':    'ב',
    'שלישי':  'ג',
    'רביעי':  'ד',
    'חמישי':  'ה',
}


def import_excel(filepath):
    """קריאה מקובץ הלוח החדש (עם עמודות יום ביקור + שבועות)"""
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb.active

    # מציאת עמודות לפי כותרת
    headers = {}
    for col in range(1, ws.max_column + 1):
        val = ws.cell(1, col).value
        if val:
            headers[str(val).strip()] = col

    RAMZOR_COL    = headers.get('רמזור', 9)
    VISIT_DAY_COL = headers.get('יום ביקור', RAMZOR_COL + 1)
    WEEK_COLS = {
        1: headers.get('שבוע 1', RAMZOR_COL + 2),
        2: headers.get('שבוע 2', RAMZOR_COL + 3),
        3: headers.get('שבוע 3', RAMZOR_COL + 4),
        4: headers.get('שבוע 4', RAMZOR_COL + 5),
    }

    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute('DELETE FROM customers')

    imported = 0
    for row in range(2, ws.max_row + 1):
        card_code    = str(ws.cell(row, 1).value or '').strip()
        name         = str(ws.cell(row, 2).value or '').strip()
        city         = str(ws.cell(row, 3).value or '').strip()
        address      = str(ws.cell(row, 4).value or '').strip()
        region       = str(ws.cell(row, 5).value or '').strip()
        delivery_day = str(ws.cell(row, 6).value or '').strip()
        x_val        = ws.cell(row, 7).value
        visit_day    = str(ws.cell(row, 8).value or '').strip()

        if not name or not region:
            continue

        # צבע רמזור
        tl_cell = ws.cell(row, RAMZOR_COL)
        traffic_light = ''
        if tl_cell.fill and tl_cell.fill.fill_type == 'solid':
            fg = tl_cell.fill.fgColor
            if fg and fg.type == 'rgb':
                traffic_light = COLOR_TO_TL.get(fg.rgb, '')

        # יום ביקור מותאם (מהלוח החדש)
        day_name = str(ws.cell(row, VISIT_DAY_COL).value or '').strip()
        assigned_visit_day = VISIT_DAY_MAP.get(day_name, visit_day)

        # שבועות (1 = מבקרים באותו שבוע)
        week_data = {}
        for wk, col in WEEK_COLS.items():
            cell_val = ws.cell(row, col).value
            week_data[wk] = 1 if cell_val else 0

        try:
            x = int(float(str(x_val))) if x_val else 0
        except Exception:
            x = 0

        c.execute('''INSERT INTO customers
            (card_code, name, city, address, region, delivery_day, x_days, visit_day,
             traffic_light, assigned_visit_day, week_1, week_2, week_3, week_4)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?)''', (
            card_code, name, city, address, region, delivery_day, x, visit_day,
            traffic_light, assigned_visit_day,
            week_data[1], week_data[2], week_data[3], week_data[4]
        ))
        imported += 1

    conn.commit()
    conn.close()
    print(f"Imported {imported} customers OK")


if __name__ == '__main__':
    filepath = sys.argv[1] if len(sys.argv) > 1 else 'סידור_עבודה_חדש.xlsx'
    import_excel(filepath)
