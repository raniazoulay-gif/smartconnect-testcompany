"""
דוח סידור עבודה חודשי לכל סוכן — מאי 2026
מריץ מקומית, שומר Excel על שולחן העבודה
"""
import sqlite3, calendar, os, subprocess
from datetime import date
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ─── הגדרות ───────────────────────────────────────────────────────────────────
DB_PATH   = os.path.join(os.path.dirname(__file__), "sales.db")
OUT_PATH  = r"C:\Users\Ran Azoulay\OneDrive - Matrix IT Ltd\Desktop\Desktop\Drive\RAN\Claude\SKILL\GOOD\GILI\קו לקוחות מעודכן\schedule_may2026.xlsx"
YEAR, MON = 2026, 5

HEBREW_DAYS = {0:'ב', 1:'ג', 2:'ד', 3:'ה', 4:'ו', 5:'ש', 6:'א'}
HEB_NAME    = {'א':'ראשון','ב':'שני','ג':'שלישי','ד':'רביעי','ה':'חמישי','ו':'שישי','ש':'שבת'}

# ─── פונקציות עזר ──────────────────────────────────────────────────────────────
def get_week_of_month(d):
    """מחזור 6 שבועות רולינג מ-11/5/2026 — זהה לפונקציה ב-main.py"""
    CYCLE_START = date(2026, 5, 11)
    days = max(0, (d - CYCLE_START).days)
    return days // 7 % 6 + 1

def border():
    s = Side(style="thin", color="B8C5E0")
    return Border(left=s, right=s, top=s, bottom=s)

def set_date_row(ws, row, text, ncols):
    """שורת כותרת תאריך — כחול כהה"""
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    c = ws.cell(row=row, column=1, value=text)
    c.font      = Font(bold=True, color="FFFFFF", size=12, name="Arial")
    c.fill      = PatternFill("solid", start_color="00327d")
    c.alignment = Alignment(horizontal="right", vertical="center")
    c.border    = border()
    ws.row_dimensions[row].height = 22

def set_customer_row(ws, row, vals, alt):
    """שורת לקוח — לסירוגין לבן/תכלת"""
    bg = "EEF1F8" if alt else "FFFFFF"
    for col, val in enumerate(vals, 1):
        c = ws.cell(row=row, column=col, value=val)
        c.font      = Font(size=10, name="Arial")
        c.fill      = PatternFill("solid", start_color=bg)
        c.alignment = Alignment(horizontal="right", vertical="center")
        c.border    = border()

# ─── מסד הנתונים ──────────────────────────────────────────────────────────────
conn = sqlite3.connect(DB_PATH)
conn.row_factory = sqlite3.Row
db = conn.cursor()

agents = db.execute(
    "SELECT id, name, regions FROM users WHERE role='agent' ORDER BY name"
).fetchall()

_, days_in = calendar.monthrange(YEAR, MON)
all_dates  = [date(YEAR, MON, d) for d in range(1, days_in + 1)]

# ─── בניית Excel ──────────────────────────────────────────────────────────────
wb = Workbook()
wb.remove(wb.active)

NCOLS = 3   # שם לקוח | עיר | אזור
COL_WIDTHS = [40, 18, 18]
HDR_VALS   = ['שם לקוח', 'עיר', 'אזור']

summary_data = []

for agent in agents:
    raw_regions = [r.strip() for r in (agent['regions'] or '').split(',') if r.strip()]
    if not raw_regions:
        continue

    ph = ','.join(['?'] * len(raw_regions))

    # קבץ כל הלקוחות ביום ↦ תאריך → [(name,city,region), ...]
    day_schedule = {}  # date → list of tuples

    for d in all_dates:
        heb = HEBREW_DAYS[d.weekday()]
        wk  = get_week_of_month(d)
        if wk > 4:
            continue  # אין week_5/week_6 בDB

        # שאילתה ראשית: assigned_visit_day
        custs = db.execute(
            f"SELECT name, city, region FROM customers"
            f" WHERE region IN ({ph}) AND assigned_visit_day=? AND week_{wk}=1"
            f" ORDER BY region, name",
            raw_regions + [heb]
        ).fetchall()

        # fallback: visit_day (אם assigned_visit_day ריק)
        if not custs:
            custs = db.execute(
                f"SELECT name, city, region FROM customers"
                f" WHERE region IN ({ph}) AND visit_day=?"
                f" ORDER BY region, name",
                raw_regions + [heb]
            ).fetchall()

        if custs:
            day_schedule[d] = [(c['name'], c['city'] or '', c['region'] or '') for c in custs]

    # ─── גיליון לסוכן ──────────────────────────────────────────────────────────
    ws = wb.create_sheet(title=agent['name'][:31])
    ws.sheet_view.rightToLeft = True

    # כותרת עליונה
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=NCOLS)
    c = ws.cell(row=1, column=1,
                value=f"UNICO  |  סידור עבודה — {agent['name']}  |  מאי {YEAR}")
    c.font      = Font(bold=True, size=14, color="FFFFFF", name="Arial")
    c.fill      = PatternFill("solid", start_color="00327d")
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border    = border()
    ws.row_dimensions[1].height = 32

    # כותרת עמודות
    for col, h in enumerate(HDR_VALS, 1):
        c = ws.cell(row=2, column=col, value=h)
        c.font      = Font(bold=True, color="FFFFFF", size=11, name="Arial")
        c.fill      = PatternFill("solid", start_color="1a4fa0")
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border    = border()
    ws.row_dimensions[2].height = 20

    cur_row = 3
    alt = False

    for d in all_dates:
        if d not in day_schedule:
            continue
        custs = day_schedule[d]
        day_label = HEB_NAME.get(HEBREW_DAYS[d.weekday()], '')
        date_text = f"{d.strftime('%d/%m/%Y')} — יום {day_label}  ({len(custs)} לקוחות)"
        set_date_row(ws, cur_row, date_text, NCOLS)
        cur_row += 1
        for (name, city, region) in custs:
            set_customer_row(ws, cur_row, [name, city, region], alt)
            alt = not alt
            cur_row += 1

    # שורת סיכום
    total_visits = sum(len(v) for v in day_schedule.values())
    work_days    = len(day_schedule)
    ws.merge_cells(start_row=cur_row+1, start_column=1, end_row=cur_row+1, end_column=NCOLS)
    c = ws.cell(row=cur_row+1, column=1,
                value=f'סה"כ ביקורים מתוכננים: {total_visits}  |  ימי עבודה: {work_days}')
    c.font      = Font(bold=True, size=11, color="00327d", name="Arial")
    c.fill      = PatternFill("solid", start_color="FFF9C4")
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border    = border()
    ws.row_dimensions[cur_row+1].height = 22

    # רוחב עמודות
    for i, w in enumerate(COL_WIDTHS):
        ws.column_dimensions[chr(65 + i)].width = w

    ws.freeze_panes = "A3"

    summary_data.append({
        'name': agent['name'],
        'visits': total_visits,
        'days': work_days,
    })

# ─── גיליון סיכום ─────────────────────────────────────────────────────────────
ws_sum = wb.create_sheet(title='סיכום', index=0)
ws_sum.sheet_view.rightToLeft = True

ws_sum.merge_cells("A1:D1")
c = ws_sum["A1"]
c.value     = f"UNICO  |  סיכום סידור עבודה — מאי {YEAR}"
c.font      = Font(bold=True, size=14, color="FFFFFF", name="Arial")
c.fill      = PatternFill("solid", start_color="00327d")
c.alignment = Alignment(horizontal="center", vertical="center")
c.border    = border()
ws_sum.row_dimensions[1].height = 32

hdrs = ['סוכן', 'ביקורים מתוכננים', 'ימי עבודה', 'ממוצע לקוחות ביום']
for col, h in enumerate(hdrs, 1):
    c = ws_sum.cell(row=2, column=col, value=h)
    c.font      = Font(bold=True, color="FFFFFF", size=11, name="Arial")
    c.fill      = PatternFill("solid", start_color="1a4fa0")
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border    = border()
ws_sum.row_dimensions[2].height = 20

for i, s in enumerate(summary_data):
    row = 3 + i
    avg = round(s['visits'] / s['days'], 1) if s['days'] else 0
    vals = [s['name'], s['visits'], s['days'], avg]
    bg  = "FFFFFF" if i % 2 == 0 else "EEF1F8"
    for col, val in enumerate(vals, 1):
        c = ws_sum.cell(row=row, column=col, value=val)
        c.font      = Font(size=11, name="Arial")
        c.fill      = PatternFill("solid", start_color=bg)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border    = border()

# שורת סיכום כללי
total_row = 3 + len(summary_data)
ws_sum.merge_cells(f"A{total_row}:C{total_row}")
c = ws_sum.cell(row=total_row, column=1,
                value=f'סה"כ ביקורים בחברה: {sum(s["visits"] for s in summary_data)}')
c.font      = Font(bold=True, size=11, color="00327d", name="Arial")
c.fill      = PatternFill("solid", start_color="FFF9C4")
c.alignment = Alignment(horizontal="center", vertical="center")
c.border    = border()
ws_sum.cell(row=total_row, column=4).fill = PatternFill("solid", start_color="FFF9C4")
ws_sum.cell(row=total_row, column=4).border = border()
ws_sum.row_dimensions[total_row].height = 22

ws_sum.column_dimensions['A'].width = 20
ws_sum.column_dimensions['B'].width = 22
ws_sum.column_dimensions['C'].width = 16
ws_sum.column_dimensions['D'].width = 22
ws_sum.freeze_panes = "A3"

conn.close()

# ─── שמירה ופתיחה ─────────────────────────────────────────────────────────────
wb.save(OUT_PATH)
print(f"✅ הקובץ נשמר: {OUT_PATH}")
os.startfile(OUT_PATH)
