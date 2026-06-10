from fastapi import FastAPI, Request, Form, Depends, HTTPException
from fastapi.responses import HTMLResponse, RedirectResponse, JSONResponse, StreamingResponse
import io
from urllib.parse import quote
from fastapi.templating import Jinja2Templates
from fastapi.staticfiles import StaticFiles
from datetime import datetime, timedelta
from zoneinfo import ZoneInfo
import json
import logging
import os
import time
from itsdangerous import TimestampSigner, BadSignature, SignatureExpired

_IL_TZ = ZoneInfo('Asia/Jerusalem')
def now_il() -> datetime:
    """datetime נוכחי בשעון ישראל, ללא tzinfo (naive) לשמירה ב-DB."""
    return datetime.now(_IL_TZ).replace(tzinfo=None)
from database import get_db, init_db, hash_password, check_password, get_setting, set_setting, DATABASE_URL, _is_postgres
from import_data import import_excel
from apscheduler.schedulers.background import BackgroundScheduler
from admin_panel import admin_router, cleanup_deleted_customers

app = FastAPI()
app.include_router(admin_router)
templates = Jinja2Templates(directory="templates")
app.mount("/static", StaticFiles(directory="static"), name="static")

import pathlib as _pathlib
_cfg_path = _pathlib.Path(__file__).parent / "company_config.json"
COMPANY_CONFIG = json.loads(_cfg_path.read_text(encoding="utf-8")) if _cfg_path.exists() else {}

# True only after startup() finishes — Railway health check blocks traffic until then
APP_READY = False
# Module-level scheduler reference so shutdown() can reach it
_scheduler: BackgroundScheduler = None
# Timestamp of when this container process started
_CONTAINER_START_TIME = time.time()
# Whether this container already cleared maintenance
_maintenance_cleared = False

MAINTENANCE_FLAG = "maintenance.flag"
ADMIN_SECRET = os.environ.get("ADMIN_SECRET", "")
COOKIE_SECRET = os.environ.get("SECRET_KEY")
if not COOKIE_SECRET:
    raise RuntimeError("SECRET_KEY env var is required — set it in Railway before deploying")
_signer = TimestampSigner(COOKIE_SECRET)

def is_maintenance():
    """בודק: env var → קובץ מקומי → DB (הכי אמין ב-Railway)"""
    if os.environ.get("MAINTENANCE_MODE") == "1":
        return True
    if os.path.exists(MAINTENANCE_FLAG):
        return True
    try:
        if get_setting("maintenance", "0") != "1":
            return False
        # Auto-reset: if maintenance was ON for more than 2 hours, clear it automatically
        maintenance_on_at = float(get_setting("maintenance_on_at", "0"))
        if maintenance_on_at > 0 and (time.time() - maintenance_on_at) > 7200:
            set_setting("maintenance", "0")
            return False
        return True
    except Exception:
        # DB unavailable (container startup) — assume maintenance for first 3 minutes
        return (time.time() - _CONTAINER_START_TIME) < 180
MAINTENANCE_HTML = """<!DOCTYPE html><html lang="he" dir="rtl"><head><meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>__COMPANY_NAME__ — עדכון</title>
<style>
*{margin:0;padding:0;box-sizing:border-box}
body{font-family:'Segoe UI',sans-serif;min-height:100vh;display:flex;align-items:center;justify-content:center;
background:radial-gradient(ellipse 80% 60% at 50% -10%,rgba(99,102,241,.35) 0%,transparent 60%),linear-gradient(180deg,#060d1f 0%,#0d1630 100%);
color:#f1f5f9;direction:rtl;text-align:center;padding:24px}
.box{background:rgba(255,255,255,.07);border:1px solid rgba(245,158,11,.4);border-radius:24px;padding:40px 32px;max-width:380px;width:100%;box-shadow:0 0 40px rgba(245,158,11,.15)}
.icon{font-size:52px;margin-bottom:18px}
h1{font-size:22px;font-weight:700;margin-bottom:10px}
p{font-size:15px;color:rgba(241,245,249,.6);line-height:1.6}
.spinner{width:36px;height:36px;border:3px solid rgba(245,158,11,.2);border-top-color:#f59e0b;border-radius:50%;animation:spin 1s linear infinite;margin:20px auto 0}
@keyframes spin{to{transform:rotate(360deg)}}
</style></head><body>
<div class="box">
  <div class="icon">🔧</div>
  <h1>האפליקציה בעדכון</h1>
  <p>אנחנו מבצעים שדרוג קצר.<br>נא להמתין כמה דקות ולנסות שוב.</p>
  <div class="spinner"></div>
</div>
<script>
(function poll(){
  fetch('/api/maintenance-status').then(function(r){return r.json();}).then(function(d){
    if(!d.maintenance){location.reload();}else{setTimeout(poll,5000);}
  }).catch(function(){setTimeout(poll,5000);});
})();
</script>
</body></html>"""

@app.get("/health")
async def health_check():
    if not APP_READY:
        return JSONResponse({"status": "starting"}, status_code=503)
    return JSONResponse({"status": "ok"})

@app.middleware("http")
async def maintenance_middleware(request: Request, call_next):
    global _maintenance_cleared
    path = request.url.path
    # Always allow health check (Railway uses this to gate traffic)
    if path in ("/health", "/healthz", "/ping"):
        return await call_next(request)
    # Block everything while app is still initializing
    if not APP_READY:
        return HTMLResponse(MAINTENANCE_HTML.replace("__COMPANY_NAME__", COMPANY_CONFIG.get("company_name", "SmartConnect")), status_code=503)
    # Clear maintenance only if this container started AFTER maintenance was enabled.
    # This prevents the old (still-running) container from accidentally clearing it.
    if not _maintenance_cleared:
        try:
            maintenance_on_at = float(get_setting("maintenance_on_at", "0"))
            now = time.time()
            container_age = now - _CONTAINER_START_TIME
            elapsed_since_on = now - maintenance_on_at
            # נקה תחזוקה רק אם:
            # 1. maintenance_on_at הוגדר בפועל (לא ברירת מחדל 0)
            # 2. container זה הופעל אחרי שהתחזוקה הופעלה (container חדש)
            # 3. עברו לפחות 3 דקות מאז שהתחזוקה הופעלה
            # 4. container זה פעיל לפחות 3 דקות (Railway deploy ~2 דקות)
            if (maintenance_on_at > 0
                    and _CONTAINER_START_TIME > maintenance_on_at
                    and elapsed_since_on > 180
                    and container_age > 180):
                _maintenance_cleared = True
                set_setting("maintenance", "0")
        except Exception:
            pass
    if is_maintenance():
        skip = (path.startswith("/api/maintenance")
                or path.startswith("/api/admin/")
                or path.startswith("/static")
                or path in ("/", "/login", "/logout"))
        if not skip:
            return HTMLResponse(MAINTENANCE_HTML.replace("__COMPANY_NAME__", COMPANY_CONFIG.get("company_name", "SmartConnect")), status_code=503)
    return await call_next(request)

@app.get("/api/maintenance-status")
async def maintenance_status():
    return JSONResponse({"maintenance": is_maintenance()})

@app.post("/api/admin/maintenance/on")
async def maintenance_on(request: Request):
    token = request.query_params.get("token", "")
    if not ADMIN_SECRET or token != ADMIN_SECRET:
        raise HTTPException(status_code=403, detail="Forbidden")
    set_setting("maintenance", "1")
    set_setting("maintenance_on_at", str(time.time()))
    return JSONResponse({"maintenance": True, "status": "on"})

@app.post("/api/admin/maintenance/off")
async def maintenance_off(request: Request):
    token = request.query_params.get("token", "")
    if not ADMIN_SECRET or token != ADMIN_SECRET:
        raise HTTPException(status_code=403, detail="Forbidden")
    set_setting("maintenance", "0")
    return JSONResponse({"maintenance": False, "status": "off"})

@app.post("/api/admin/update-user")
async def admin_update_user(
    request: Request,
    old_username: str = Form(...),
    new_username: str = Form(...),
    new_password: str = Form(...)
):
    token = request.query_params.get("token", "")
    if not ADMIN_SECRET or token != ADMIN_SECRET:
        raise HTTPException(status_code=403, detail="Forbidden")
    db = get_db()
    try:
        db.execute(
            "UPDATE users SET username=?, password_hash=? WHERE username=?",
            (new_username, hash_password(new_password), old_username)
        )
        db.commit()
        return JSONResponse({"ok": True, "updated": new_username})
    finally:
        db.close()

@app.post("/api/admin/sync-customers")
async def admin_sync_customers(request: Request):
    token = request.query_params.get("token", "")
    if not ADMIN_SECRET or token != ADMIN_SECRET:
        raise HTTPException(status_code=403, detail="Forbidden")
    body = await request.json()
    customers = body.get("customers", [])
    db = get_db()
    inserted = updated = 0
    try:
        for c in customers:
            existing = db.execute(
                "SELECT id FROM customers WHERE card_code=?", (c["card_code"],)
            ).fetchone()
            if existing:
                db.execute('''UPDATE customers SET name=?,city=?,region=?,visit_day=?,
                    traffic_light=?,week_1=?,week_2=?,week_3=?,week_4=? WHERE card_code=?''',
                    (c["name"],c["city"],c["region"],c["visit_day"],
                     c["traffic_light"],c["week_1"],c["week_2"],c["week_3"],c["week_4"],
                     c["card_code"]))
                updated += 1
            else:
                db.execute('''INSERT INTO customers
                    (card_code,name,city,region,visit_day,traffic_light,week_1,week_2,week_3,week_4)
                    VALUES (?,?,?,?,?,?,?,?,?,?)''',
                    (c["card_code"],c["name"],c["city"],c["region"],c["visit_day"],
                     c["traffic_light"],c["week_1"],c["week_2"],c["week_3"],c["week_4"]))
                inserted += 1
        db.commit()
        return JSONResponse({"ok": True, "inserted": inserted, "updated": updated})
    finally:
        db.close()

@app.post("/api/admin/sync-user-regions")
async def admin_sync_user_regions(request: Request):
    token = request.query_params.get("token", "")
    if not ADMIN_SECRET or token != ADMIN_SECRET:
        raise HTTPException(status_code=403, detail="Forbidden")
    body = await request.json()
    db = get_db()
    try:
        for username, regions in body.items():
            db.execute("UPDATE users SET regions=? WHERE username=?", (regions, username))
        db.commit()
        return JSONResponse({"ok": True})
    finally:
        db.close()

@app.get("/api/admin/debug-user")
async def admin_debug_user(request: Request, username: str = ""):
    token = request.query_params.get("token", "")
    if not ADMIN_SECRET or token != ADMIN_SECRET:
        raise HTTPException(status_code=403, detail="Forbidden")
    db = get_db()
    try:
        user = db.execute("SELECT id,name,username,regions FROM users WHERE username=?", (username,)).fetchone()
        if not user:
            return JSONResponse({"error": "user not found"})
        regions = [r.strip() for r in user["regions"].split(",") if r.strip()]
        counts = {}
        for r in regions:
            cnt = db.execute("SELECT COUNT(*) FROM customers WHERE region=?", (r,)).fetchone()[0]
            counts[r] = cnt
        return JSONResponse({"user": {k: user[k] for k in user.keys()}, "region_counts": counts, "total": sum(counts.values())})
    finally:
        db.close()

HEBREW_DAYS = {0: 'ב', 1: 'ג', 2: 'ד', 3: 'ה', 4: 'ו', 5: 'ש', 6: 'א'}
HEBREW_DAY_NAMES = {'א': 'ראשון', 'ב': 'שני', 'ג': 'שלישי', 'ד': 'רביעי', 'ה': 'חמישי', 'ו': 'שישי', 'ש': 'שבת'}
TRAFFIC_LIGHT_COLORS = {'ירוק': '#92D050', 'כתום': '#FFC000', 'אדום': '#FF0000'}

def get_hebrew_day(date_obj):
    return HEBREW_DAYS[date_obj.weekday()]

def get_week_of_month(date_obj, cycle=6):
    """
    cycle=7 (גילי): מחזור רולינג 7 שבועות מ-1/5/2026
      ש1=1-7/5, ש2=8-14/5, ..., ש7=12-18/6, חוזר לש1
    cycle=4 (שיראל): שבוע 1 תמיד מתחיל ב-1 לחודש
      ש1=ימים 1-7, ש2=8-14, ש3=15-21, ש4=22+
    cycle=6 (שאר): מחזור רולינג מ-11/5/2026
    """
    from datetime import date as _date
    if cycle == 7:
        GILI_CYCLE_START = _date(2026, 5, 1)
        days = max(0, (date_obj - GILI_CYCLE_START).days)
        return days // 7 % 7 + 1
    if cycle == 4:
        return min(4, (date_obj.day - 1) // 7 + 1)
    CYCLE_START = _date(2026, 5, 11)
    days = max(0, (date_obj - CYCLE_START).days)
    return days // 7 % cycle + 1

def get_agent_cycle(user):
    """מחזיר cycle לפי סוכן (ברירת מחדל, ללא ידיעת יום)"""
    if not user:
        return 6
    try:
        name     = user['name']     or ''
        username = user['username'] or ''
    except (KeyError, TypeError):
        return 6
    if name == 'גילי' or username == 'gili':
        return 7   # ברירת מחדל לגילי (רביעי) — override בפונקציה get_week_for_agent
    if name == 'שיראל' or username == 'shirel':
        return 4
    return 6

def get_week_for_agent(date_obj, user):
    """
    מחזיר מספר שבוע נכון לפי סוכן + תאריך + יום שבוע.
    גילי — רביעי: cycle 7 רולינג מ-1/5/2026
    גילי — שני: cycle 5 רולינג מ-1/5/2026
    גילי — שאר ימים: cycle 4 רולינג מ-1/5/2026
    שיראל / אלי: cycle 4 לפי יום בחודש (1-7=ש1, 8-14=ש2, 15-21=ש3, 22+=ש4)
    שאר: cycle 6 רולינג מ-11/5/2026
    """
    from datetime import date as _date
    if not user:
        CYCLE_START = _date(2026, 5, 11)
        return max(0, (date_obj - CYCLE_START).days) // 7 % 6 + 1
    try:
        name     = user['name']     or ''
        username = user['username'] or ''
    except (KeyError, TypeError):
        name = username = ''

    if name == 'גילי' or username == 'gili':
        GILI_START = _date(2026, 5, 1)
        days = max(0, (date_obj - GILI_START).days)
        if date_obj.weekday() == 2:    # רביעי = Wednesday → 7 שבועות
            return days // 7 % 7 + 1
        elif date_obj.weekday() == 0:  # שני = Monday → 5 שבועות
            return days // 7 % 5 + 1
        else:                          # שאר הימים → 4 שבועות
            return days // 7 % 4 + 1

    if name == 'שיראל' or username == 'shirel':
        return min(4, (date_obj.day - 1) // 7 + 1)

    if name == 'אלי' or username == 'eli':
        return min(4, (date_obj.day - 1) // 7 + 1)

    CYCLE_START = _date(2026, 5, 11)
    return max(0, (date_obj - CYCLE_START).days) // 7 % 6 + 1

def get_current_user(request: Request):
    cookie = request.cookies.get("user")
    if not cookie:
        return None
    try:
        unsigned = _signer.unsign(cookie, max_age=86400 * 7).decode()
        return json.loads(unsigned)
    except (BadSignature, SignatureExpired):
        return None
    except Exception as e:
        logging.warning("get_current_user: unexpected error parsing cookie: %s", e)
        return None

def send_eod_report():
    """דוח סוף יום — נשלח ב-18:00. מזהה חנויות שלא בוקרו ויוצר התראות למנהל."""
    today = str(now_il().date())
    db = get_db()
    try:
        # כל הסוכנים
        agents = db.execute("SELECT * FROM users WHERE role='agent'").fetchall()
        hebrew_day = HEBREW_DAYS[now_il().weekday()]

        for agent in agents:
            week_num   = get_week_for_agent(now_il().date(), agent)
            week_col   = f"week_{week_num}"
            regions = [r.strip() for r in agent['regions'].split(',') if r.strip()]
            if not regions:
                continue

            placeholders = ','.join(['?' for _ in regions])

            # חנויות היום של הסוכן
            sql_stores_today = (
                "SELECT id, name FROM customers"
                " WHERE region IN (" + placeholders + ")"
                " AND assigned_visit_day=?"
                " AND " + week_col + "=1"
            )
            stores_today = db.execute(
                sql_stores_today,
                regions + [hebrew_day]
            ).fetchall()

            if not stores_today:
                continue

            # ביקורים שבוצעו (יש check_out)
            visited_ids = set()
            partial_ids = set()
            for s in stores_today:
                visit = db.execute(
                    "SELECT check_in_time, check_out_time FROM visits WHERE user_id=? AND customer_id=? AND visit_date=?",
                    (agent['id'], s['id'], today)
                ).fetchone()
                if visit:
                    if visit['check_out_time']:
                        visited_ids.add(s['id'])
                    elif visit['check_in_time']:
                        partial_ids.add(s['id'])

            missed = [s for s in stores_today if s['id'] not in visited_ids]
            total  = len(stores_today)
            done   = len(visited_ids)
            inside = len(partial_ids)

            # בנה הודעת סטטוס
            lines = [f"📊 דוח יום {today} — {agent['name']}"]
            lines.append(f"✅ הושלמו: {done}/{total}")
            if inside:
                lines.append(f"🟡 עדיין בפנים: {inside}")
            # רשימת כל החנויות — מבוקרות ❌ וגם ✓
            for s in stores_today:
                if s['id'] in visited_ids:
                    lines.append(f"   ✓ {s['name']}")
                elif s['id'] in partial_ids:
                    lines.append(f"   ⏳ {s['name']}")
                else:
                    lines.append(f"   • {s['name']}")

            msg = '\n'.join(lines)

            db.execute(
                """INSERT INTO notifications (message, user_name, store_name, action, created_at, is_read)
                   VALUES (?, ?, ?, ?, ?, 0)""",
                (msg, agent['name'], f"דוח יומי — {done}/{total}", 'eod_report', now_il().isoformat())
            )

        db.commit()
        print(f"[EOD] דוח סוף יום נשלח — {now_il()}")
    finally:
        db.close()


# ─── PUSH NOTIFICATIONS ──────────────────────────────────────────────────────

def _ensure_vapid_keys():
    """Generate VAPID keys stored as base64url raw bytes — compatible with Apple APNs"""
    import traceback
    try:
        existing_priv = get_setting("vapid_private_key", "")
        existing_pub  = get_setting("vapid_public_key",  "")

        import base64
        from cryptography.hazmat.primitives import serialization
        from cryptography.hazmat.primitives.asymmetric import ec

        # ── Case 1: already base64url raw format (no "-----") ──────────────────
        if existing_priv and existing_pub and not existing_priv.strip().startswith("-----"):
            print(f"[VAPID] Keys OK (base64url). Public: {existing_pub[:20]}...")
            return

        # ── Case 2: key exists in PEM format → convert, keep subscriptions ────
        if existing_priv and existing_priv.strip().startswith("-----"):
            print("[VAPID] Converting PEM key to base64url (no subscription change)")
            from py_vapid import Vapid
            v = Vapid.from_pem(existing_priv.strip().encode())
            # Get raw private scalar (32 bytes big-endian)
            raw_priv = v._private_key.private_numbers().private_value.to_bytes(32, 'big')
            private_b64 = base64.urlsafe_b64encode(raw_priv).rstrip(b'=').decode()
            # Recompute public key from the loaded private key (guaranteed named-curve)
            pub_bytes = v._private_key.public_key().public_bytes(
                encoding=serialization.Encoding.X962,
                format=serialization.PublicFormat.UncompressedPoint
            )
            public_b64 = base64.urlsafe_b64encode(pub_bytes).rstrip(b'=').decode()
            set_setting("vapid_private_key", private_b64)
            set_setting("vapid_public_key",  public_b64)
            print(f"[VAPID] Converted. Public: {public_b64[:20]}...")
            return

        # ── Case 3: no key at all → generate fresh ────────────────────────────
        print("[VAPID] Generating new VAPID keys")
        try:
            db = get_db()
            db.execute("DELETE FROM push_subscriptions")
            db.commit()
            db.close()
        except Exception:
            pass
        private_key_obj = ec.generate_private_key(ec.SECP256R1())
        raw_priv = private_key_obj.private_numbers().private_value.to_bytes(32, 'big')
        private_b64 = base64.urlsafe_b64encode(raw_priv).rstrip(b'=').decode()
        pub_bytes = private_key_obj.public_key().public_bytes(
            encoding=serialization.Encoding.X962,
            format=serialization.PublicFormat.UncompressedPoint
        )
        public_b64 = base64.urlsafe_b64encode(pub_bytes).rstrip(b'=').decode()
        set_setting("vapid_private_key", private_b64)
        set_setting("vapid_public_key",  public_b64)
        print(f"[VAPID] Generated. Public: {public_b64[:20]}...")
    except Exception as e:
        print(f"[VAPID] Error: {e}\n{traceback.format_exc()}")

def _send_push_to_managers(title: str, body: str, url: str = "/manager"):
    """Send web push notification to all subscribed manager devices"""
    import traceback
    # שמור פעולה ממתינה בשרת — המנהל יקרא אותה בפתיחת האפליקציה
    try:
        set_setting("manager_pending_push", json.dumps({"url": url, "ts": now_il().isoformat()}))
    except Exception:
        pass
    try:
        from pywebpush import webpush, WebPushException
    except Exception as e:
        print(f"[PUSH] pywebpush import failed: {e}")
        return
    try:
        private_key = get_setting("vapid_private_key", "")
        if not private_key:
            print("[PUSH] No VAPID private key in DB")
            return
        db = get_db()
        subs = db.execute("SELECT * FROM push_subscriptions").fetchall()
        db.close()
        print(f"[PUSH] Sending to {len(subs)} subscriptions")
        if not subs:
            return
        payload = json.dumps({"title": title, "body": body, "url": url})
        for sub in subs:
            try:
                webpush(
                    subscription_info={
                        "endpoint": sub["endpoint"],
                        "keys": {"p256dh": sub["p256dh"], "auth": sub["auth"]}
                    },
                    data=payload,
                    vapid_private_key=private_key,
                    vapid_claims={"sub": f"mailto:{COMPANY_CONFIG.get('manager_email', 'admin@smartconnect.app')}"}
                )
                print(f"[PUSH] Sent OK to {sub['endpoint'][:50]}")
            except Exception as e:
                print(f"[PUSH] Failed: {e}\n{traceback.format_exc()}")
    except Exception as e:
        print(f"[PUSH] send_push error: {e}\n{traceback.format_exc()}")

@app.get("/api/push/vapid-public-key")
async def vapid_public_key():
    key = get_setting("vapid_public_key", "")
    return JSONResponse({"publicKey": key})

@app.get("/api/manager-pending-action")
async def manager_pending_action(request: Request):
    """מחזיר ומוחק פעולה ממתינה למנהל (לפתיחת צ'אט נכון אחרי לחיצה על נוטיפיקציה)"""
    user = get_current_user(request)
    if not user:
        return JSONResponse({"error": "לא מחובר"}, status_code=401)
    if user['role'] != 'manager':
        return JSONResponse({"error": "אין הרשאה"}, status_code=403)
    raw = get_setting("manager_pending_push", "")
    if not raw:
        return JSONResponse({"action": None})
    try:
        data = json.loads(raw)
        # תקף רק ב-60 שניות האחרונות
        ts = datetime.fromisoformat(data.get("ts", "2000-01-01"))
        age = (now_il() - ts).total_seconds()
        if age > 60:
            set_setting("manager_pending_push", "")
            return JSONResponse({"action": None})
        set_setting("manager_pending_push", "")
        return JSONResponse({"action": data.get("url", "")})
    except Exception:
        set_setting("manager_pending_push", "")
        return JSONResponse({"action": None})

@app.post("/api/push/subscribe")
async def push_subscribe(request: Request):
    user = get_current_user(request)
    if not user or user['role'] != 'manager':
        return JSONResponse({"error": "אין הרשאה"}, status_code=403)
    data = await request.json()
    endpoint = data.get("endpoint", "")
    keys = data.get("keys", {})
    p256dh = keys.get("p256dh", "")
    auth   = keys.get("auth", "")
    if not endpoint or not p256dh or not auth:
        return JSONResponse({"error": "נתונים חסרים"}, status_code=400)
    db = get_db()
    try:
        existing = db.execute("SELECT id FROM push_subscriptions WHERE endpoint=?", (endpoint,)).fetchone()
        if existing:
            db.execute("UPDATE push_subscriptions SET p256dh=?, auth=?, created_at=? WHERE endpoint=?",
                       (p256dh, auth, now_il().isoformat(), endpoint))
        else:
            db.execute("INSERT INTO push_subscriptions (user_id, endpoint, p256dh, auth, created_at) VALUES (?,?,?,?,?)",
                       (user['id'], endpoint, p256dh, auth, now_il().isoformat()))
        db.commit()
    finally:
        db.close()
    print(f"[PUSH] Subscription saved for manager {user['name']} — endpoint: {endpoint[:60]}")
    return JSONResponse({"status": "ok"})

@app.delete("/api/push/unsubscribe")
async def push_unsubscribe(request: Request):
    user = get_current_user(request)
    if not user:
        return JSONResponse({"error": "לא מחובר"}, status_code=401)
    data = await request.json()
    endpoint = data.get("endpoint", "")
    db = get_db()
    db.execute("DELETE FROM push_subscriptions WHERE endpoint=?", (endpoint,))
    db.commit()
    db.close()
    return JSONResponse({"status": "ok"})

@app.get("/api/push/debug")
async def push_debug(request: Request):
    """Debug endpoint — check push setup status"""
    import traceback
    token = request.query_params.get("token", "")
    if not ADMIN_SECRET or token != ADMIN_SECRET:
        raise HTTPException(status_code=403)
    result = {}
    # VAPID keys
    try:
        pub  = get_setting("vapid_public_key", "")
        priv = get_setting("vapid_private_key", "")
        result["vapid_public_key"]  = (pub[:30]  + "...") if pub  else "MISSING"
        result["vapid_private_key"] = "EXISTS" if priv else "MISSING"
        result["vapid_priv_format"] = priv.split('\n')[0] if priv else "NONE"
    except Exception as e:
        result["vapid_error"] = str(e)
    # Subscriptions
    try:
        db = get_db()
        # Create table if missing
        db.execute('''CREATE TABLE IF NOT EXISTS push_subscriptions (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER NOT NULL,
            endpoint TEXT NOT NULL,
            p256dh TEXT NOT NULL,
            auth TEXT NOT NULL,
            created_at TEXT NOT NULL
        )''')
        db.commit()
        subs = db.execute("SELECT id, user_id, endpoint, created_at FROM push_subscriptions").fetchall()
        db.close()
        result["subscriptions"]      = [{"id": s["id"], "endpoint": s["endpoint"][:80], "created": s["created_at"]} for s in subs]
        result["subscription_count"] = len(subs)
    except Exception as e:
        result["subscriptions_error"] = f"{e}\n{traceback.format_exc()}"
    # pywebpush import test
    try:
        from pywebpush import webpush
        import pywebpush as _pw
        result["pywebpush"] = f"OK (v{getattr(_pw,'__version__','?')})"
    except Exception as e:
        result["pywebpush"] = f"IMPORT ERROR: {e}"
    return JSONResponse(result)

@app.post("/api/push/test")
async def push_test(request: Request):
    """Send a test push to all subscriptions"""
    user = get_current_user(request)
    if not user or user['role'] != 'manager':
        return JSONResponse({"error": "אין הרשאה"}, status_code=403)
    import traceback
    errors = []
    sent = 0
    try:
        from pywebpush import webpush, WebPushException
        private_key = get_setting("vapid_private_key", "")
        db = get_db()
        subs = db.execute("SELECT * FROM push_subscriptions").fetchall()
        db.close()
        if not subs:
            return JSONResponse({"ok": False, "error": "אין subscriptions — לחץ 'הפעל התראות' קודם"})
        payload = json.dumps({"title": f"🔔 בדיקת {COMPANY_CONFIG.get('company_name', 'SmartConnect')}", "body": "ההתראות עובדות!", "url": "/manager"})
        for sub in subs:
            try:
                webpush(
                    subscription_info={"endpoint": sub["endpoint"], "keys": {"p256dh": sub["p256dh"], "auth": sub["auth"]}},
                    data=payload,
                    vapid_private_key=private_key,
                    vapid_claims={"sub": f"mailto:{COMPANY_CONFIG.get('manager_email', 'admin@smartconnect.app')}"}
                )
                sent += 1
            except Exception as e:
                errors.append(f"{sub['endpoint'][:40]}: {str(e)}")
    except Exception as e:
        errors.append(f"GLOBAL: {traceback.format_exc()}")
    return JSONResponse({"ok": sent > 0, "sent": sent, "errors": errors})


@app.on_event("startup")
async def startup():
    init_db()
    if not _is_postgres():
        # Migration: columns added after initial deploy — each gets its own connection
        # so a failed migration (column already exists) doesn't abort the others
        for _table, _col, _coldef in [
            ('visits',         'paused_at',              'TEXT'),
            ('visits',         'pause_duration_minutes',  'INTEGER DEFAULT 0'),
            ('visit_comments', 'is_read',                 'INTEGER DEFAULT 0'),
            ('visits',         'is_phone',                'INTEGER DEFAULT 0'),
        ]:
            try:
                db = get_db()
                db.execute(f'ALTER TABLE {_table} ADD COLUMN {_col} {_coldef}')
                db.commit()
                db.close()
            except Exception:
                try: db.close()
                except: pass
        # Migration: push_subscriptions table
        try:
            db = get_db()
            db.execute('''CREATE TABLE IF NOT EXISTS push_subscriptions (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id INTEGER NOT NULL,
                endpoint TEXT NOT NULL,
                p256dh TEXT NOT NULL,
                auth TEXT NOT NULL,
                created_at TEXT NOT NULL
            )''')
            db.commit()
            db.close()
        except Exception:
            pass
        _ensure_vapid_keys()
        # Migration: add beep_sound column to users
        try:
            db = get_db()
            db.execute("ALTER TABLE users ADD COLUMN beep_sound INTEGER DEFAULT 4")
            db.commit()
            db.close()
        except Exception:
            try: db.close()
            except: pass
    else:
        _ensure_vapid_keys()
    # Migration: copy visit_day → assigned_visit_day for customers missing it
    # DISABLED: this migration ran once (2026-05-28). Re-enabling would overwrite
    # intentionally-unassigned customers with stale visit_day values.
    # try:
    #     db = get_db()
    #     db.execute("""
    #         UPDATE customers
    #         SET assigned_visit_day = visit_day
    #         WHERE (assigned_visit_day IS NULL OR assigned_visit_day = '')
    #           AND visit_day IS NOT NULL AND visit_day != ''
    #     """)
    #     db.commit()
    #     db.close()
    # except Exception:
    #     try: db.close()
    #     except: pass
    # Fix: zero out any negative duration_minutes stored due to clock errors
    try:
        db = get_db()
        db.execute("UPDATE visits SET duration_minutes = 0 WHERE duration_minutes < 0")
        db.commit()
        db.close()
    except Exception:
        try: db.close()
        except: pass
    # קובץ לוח עבודה חדש (עם רמזור + שבועות)
    excel_path = "סידור_עבודה_חדש.xlsx"
    if not os.path.exists(excel_path):
        excel_path = "מיפוי_לקוחות_לפי_אזור_גרסה_7.xlsx"
    if os.path.exists(excel_path):
        db = get_db()
        count = db.execute("SELECT COUNT(*) FROM customers").fetchone()[0]
        db.close()
        if count == 0:
            import_excel(excel_path)

    # Migration: Admin Panel — deleted_at + deleted_backup columns
    for _col, _coldef in [('deleted_at', 'TEXT DEFAULT NULL'), ('deleted_backup', 'TEXT DEFAULT NULL')]:
        try:
            db = get_db()
            db.execute(f'ALTER TABLE customers ADD COLUMN {_col} {_coldef}')
            db.commit()
            db.close()
        except Exception:
            try: db.close()
            except: pass

    # מתזמן — דוח ב-18:00 כל יום + ניקוי לקוחות מחוקים ב-03:00
    global _scheduler
    _scheduler = BackgroundScheduler()
    _scheduler.add_job(send_eod_report, 'cron', hour=18, minute=0, timezone='Asia/Jerusalem')
    _scheduler.add_job(cleanup_deleted_customers, 'cron', hour=3, minute=0, timezone='Asia/Jerusalem')
    _scheduler.start()
    # Signal that the app is fully ready — Railway will now route traffic here.
    # Maintenance is cleared only on the first real request (see middleware).
    if COOKIE_SECRET == "dev-key-change-in-production":
        if DATABASE_URL:
            raise RuntimeError("SECRET_KEY חייב להיות מוגדר בפרודקשיין — הגדר אותו ב-Railway Environment Variables")
        print("WARNING: SECRET_KEY env var is not set — using insecure default key!")
    global APP_READY
    APP_READY = True

@app.on_event("shutdown")
async def shutdown():
    if _scheduler and _scheduler.running:
        _scheduler.shutdown(wait=False)

# ─── AUTH ───────────────────────────────────────────────────────────────────

@app.get("/", response_class=HTMLResponse)
async def login_page(request: Request):
    user = get_current_user(request)
    if user:
        return RedirectResponse("/dashboard" if user['role'] == 'agent' else "/manager")
    return templates.TemplateResponse("login.html", {"request": request, "cfg": COMPANY_CONFIG})

@app.post("/login")
async def login(request: Request, username: str = Form(...), password: str = Form(...)):
    db = get_db()
    user = db.execute("SELECT * FROM users WHERE username=?", (username,)).fetchone()

    if not user or not check_password(password, user["password_hash"]):
        db.close()
        return templates.TemplateResponse("login.html", {"request": request, "error": "שם משתמש או סיסמה שגויים", "cfg": COMPANY_CONFIG})

    # מיגרציה שקטה: אם ה-hash ישן (SHA-256) — עדכן ל-bcrypt
    if not user["password_hash"].startswith("$2"):
        db.execute("UPDATE users SET password_hash=? WHERE id=?",
                   (hash_password(password), user["id"]))
        db.commit()
    db.close()

    user_data = json.dumps({"id": user["id"], "name": user["name"], "username": user["username"],
                             "role": user["role"], "regions": user["regions"]})
    signed = _signer.sign(user_data.encode()).decode()
    redirect_url = "/dashboard" if user["role"] == "agent" else "/manager"
    response = RedirectResponse(redirect_url, status_code=302)
    response.set_cookie("user", signed, max_age=86400 * 7, httponly=True, samesite="lax", secure=True)
    return response

@app.get("/logout")
async def logout():
    response = RedirectResponse("/")
    response.delete_cookie("user")
    return response

# ─── AGENT DASHBOARD ────────────────────────────────────────────────────────

@app.get("/dashboard", response_class=HTMLResponse)
async def agent_dashboard(request: Request, date: str = None):
    user = get_current_user(request)
    if not user:
        return RedirectResponse("/")
    if user['role'] == 'manager':
        return RedirectResponse("/manager")

    try:
        selected_date = datetime.strptime(date, "%Y-%m-%d").date() if date else now_il().date()
    except ValueError:
        selected_date = now_il().date()

    hebrew_day = get_hebrew_day(selected_date)
    week_num   = get_week_for_agent(selected_date, user)
    week_col   = f"week_{week_num}"
    regions = [r.strip() for r in user['regions'].split(',') if r.strip()] or ['__none__']

    db = get_db()
    placeholders = ','.join(['?' for _ in regions])

    # פלטר לפי: אזור + יום ביקור מותאם + שבוע בחודש
    sql_customers = (
        "SELECT * FROM customers"
        " WHERE region IN (" + placeholders + ")"
        " AND assigned_visit_day=?"
        " AND " + week_col + "=1"
        " ORDER BY region, name"
    )
    customers = db.execute(
        sql_customers,
        regions + [hebrew_day]
    ).fetchall()

    # אם אין תוצאות (לפני ייבוא הנתונים החדשים) — fallback לשדה הישן
    if not customers:
        sql_customers_fallback = (
            "SELECT * FROM customers WHERE region IN (" + placeholders + ") AND visit_day=? ORDER BY region, name"
        )
        customers = db.execute(
            sql_customers_fallback,
            regions + [hebrew_day]
        ).fetchall()

    # Get visit status for each customer
    visits_today = db.execute(
        "SELECT * FROM visits WHERE user_id=? AND visit_date=?",
        (user['id'], str(selected_date))
    ).fetchall()

    eilat_customers = [dict(c) for c in db.execute(
        "SELECT * FROM customers WHERE region='שיראל-אילת' ORDER BY name"
    ).fetchall()] if user['username'] in ('shirel', 'shirael') else []

    db.close()

    visits_map = {v['customer_id']: dict(v) for v in visits_today}

    prev_date = (selected_date - timedelta(days=1)).strftime("%Y-%m-%d")
    next_date = (selected_date + timedelta(days=1)).strftime("%Y-%m-%d")

    # Build full month calendar grid (starts Sunday)
    import calendar
    year, month = selected_date.year, selected_date.month
    first_day = selected_date.replace(day=1)
    # Python weekday: Mon=0...Sun=6 → Israeli start Sunday
    # offset: how many empty cells before day 1 (Sunday=0 offset)
    first_weekday = (first_day.weekday() + 1) % 7  # 0=Sun,1=Mon...6=Sat
    days_in_month = calendar.monthrange(year, month)[1]
    month_days = []  # list of date or None for padding
    for _ in range(first_weekday):
        month_days.append(None)
    for d in range(1, days_in_month + 1):
        month_days.append(selected_date.replace(day=d))
    # pad to complete last row
    while len(month_days) % 7 != 0:
        month_days.append(None)

    month_name_map = {1:'ינואר',2:'פברואר',3:'מרץ',4:'אפריל',5:'מאי',6:'יוני',
                      7:'יולי',8:'אוגוסט',9:'ספטמבר',10:'אוקטובר',11:'נובמבר',12:'דצמבר'}
    month_label = f"{month_name_map.get(month,'')} {year}"
    week_days = []  # keep for compatibility

    return templates.TemplateResponse("agent.html", {
        "request": request,
        "user": user,
        "customers": [dict(c) for c in customers],
        "visits_map": visits_map,
        "selected_date": selected_date,
        "selected_date_str": str(selected_date),
        "hebrew_day": hebrew_day,
        "hebrew_day_name": HEBREW_DAY_NAMES.get(hebrew_day, ''),
        "week_num": week_num,
        "prev_date": prev_date,
        "next_date": next_date,
        "week_days": week_days,
        "month_days": month_days,
        "month_label": month_label,
        "timedelta": timedelta,
        "today": now_il().date(),
        "is_after_17": now_il().hour >= 17,
        "HEBREW_DAYS": HEBREW_DAYS,
        "TRAFFIC_LIGHT_COLORS": TRAFFIC_LIGHT_COLORS,
        "eilat_customers": eilat_customers,
        "cfg": COMPANY_CONFIG,
    })

# ─── STORE DETAIL ────────────────────────────────────────────────────────────

@app.get("/store/{customer_id}", response_class=HTMLResponse)
async def store_detail(request: Request, customer_id: int, date: str = None, agent_id: int = None):
    user = get_current_user(request)
    if not user:
        return RedirectResponse("/")

    try:
        visit_date = datetime.strptime(date, "%Y-%m-%d").date() if date else now_il().date()
    except ValueError:
        visit_date = now_il().date()

    is_manager = (user['role'] == 'manager')

    db = get_db()
    customer = db.execute("SELECT * FROM customers WHERE id=?", (customer_id,)).fetchone()
    if not customer:
        db.close()
        raise HTTPException(404)

    # אם מנהל — מצא את הביקור של הסוכן הרלוונטי
    if is_manager and agent_id:
        visit = db.execute(
            "SELECT * FROM visits WHERE user_id=? AND customer_id=? AND visit_date=?",
            (agent_id, customer_id, str(visit_date))
        ).fetchone()
        back_url = f"/manager/agent/{agent_id}?date={visit_date}"
    elif is_manager:
        # כניסה ממנהל ללא agent_id (לדוג' חיפוש לקוח) — הצג ביקור אחרון של כל סוכן היום
        visit = db.execute(
            "SELECT * FROM visits WHERE customer_id=? AND visit_date=? ORDER BY check_in_time DESC",
            (customer_id, str(visit_date))
        ).fetchone()
        back_url = f"/manager?date={visit_date}"
    else:
        visit = db.execute(
            "SELECT * FROM visits WHERE user_id=? AND customer_id=? AND visit_date=?",
            (user['id'], customer_id, str(visit_date))
        ).fetchone()
        back_url = f"/dashboard?date={visit_date}"

    if is_manager:
        history = db.execute(
            """SELECT v.*, u.name as agent_name FROM visits v
               JOIN users u ON v.user_id = u.id
               WHERE v.customer_id=?
               ORDER BY v.visit_date DESC, v.check_in_time DESC""",
            (customer_id,)
        ).fetchall()
    else:
        history = db.execute(
            """SELECT v.*, u.name as agent_name FROM visits v
               JOIN users u ON v.user_id = u.id
               WHERE v.customer_id=? AND v.user_id=?
               ORDER BY v.visit_date DESC, v.check_in_time DESC""",
            (customer_id, user['id'])
        ).fetchall()
    db.close()

    return templates.TemplateResponse("store.html", {
        "request": request,
        "user": user,
        "customer": dict(customer),
        "visit": dict(visit) if visit else None,
        "history": [dict(h) for h in history],
        "visit_date": str(visit_date),
        "is_manager": is_manager,
        "back_url": back_url,
        "HEBREW_DAY_NAMES": HEBREW_DAY_NAMES,
        "cfg": COMPANY_CONFIG,
    })

# ─── VISIT API ───────────────────────────────────────────────────────────────

@app.post("/api/checkin")
async def checkin(request: Request):  # v2
    user = get_current_user(request)
    if not user:
        return JSONResponse({"error": "לא מחובר"}, status_code=401)

    data = await request.json()
    customer_id = data.get("customer_id")
    if not customer_id:
        return JSONResponse({"error": "customer_id חסר"}, status_code=400)
    visit_date = data.get("visit_date", str(now_il().date()))
    now = now_il().strftime("%H:%M:%S")

    db = get_db()
    existing = db.execute(
        "SELECT id, check_in_time FROM visits WHERE user_id=? AND customer_id=? AND visit_date=?",
        (user['id'], customer_id, visit_date)
    ).fetchone()

    if existing and existing['check_in_time']:
        db.close()
        return JSONResponse({"error": "כבר נכנסת לחנות זו היום"})

    # בדיקה אם יש ביקור פתוח בחנות אחרת
    open_visit_simple = db.execute(
        """SELECT id, customer_id FROM visits
           WHERE user_id=? AND visit_date=? AND check_in_time IS NOT NULL AND check_out_time IS NULL""",
        (user['id'], visit_date)
    ).fetchone()
    if open_visit_simple:
        open_cust = db.execute("SELECT name FROM customers WHERE id=?", (open_visit_simple['customer_id'],)).fetchone()
        store_name_blocked = open_cust['name'] if open_cust else f"חנות #{open_visit_simple['customer_id']}"
        db.close()
        return JSONResponse({
            "error": f"יש לסיים ביקור בחנות {store_name_blocked} לפני כניסה לחנות חדשה",
            "open_store_name": store_name_blocked,
            "open_customer_id": open_visit_simple['customer_id'],
            "visit_date": visit_date
        })

    if existing:
        # רשומה קיימת ללא check_in_time (הערות נשמרו לפני כניסה) — עדכן check_in_time
        db.execute(
            "UPDATE visits SET check_in_time=? WHERE id=?",
            (now, existing['id'])
        )
    else:
        db.execute(
            "INSERT INTO visits (user_id, customer_id, visit_date, check_in_time) VALUES (?,?,?,?)",
            (user['id'], customer_id, visit_date, now)
        )

    customer = db.execute("SELECT name FROM customers WHERE id=?", (customer_id,)).fetchone()
    store_name = customer['name'] if customer else str(customer_id)

    db.execute(
        "INSERT INTO notifications (message, user_name, store_name, action, created_at) VALUES (?,?,?,?,?)",
        (f"{user['name']} נכנס ל{store_name}", user['name'], store_name, 'checkin', now_il().isoformat())
    )
    db.commit()
    db.close()

    return JSONResponse({"status": "ok", "time": now})

@app.post("/api/checkout")
async def checkout(request: Request):
    user = get_current_user(request)
    if not user:
        return JSONResponse({"error": "לא מחובר"}, status_code=401)

    data = await request.json()
    customer_id = data.get("customer_id")
    visit_date = data.get("visit_date", str(now_il().date()))
    notes = data.get("notes", "")
    now = now_il().strftime("%H:%M:%S")

    db = get_db()
    visit = db.execute(
        "SELECT * FROM visits WHERE user_id=? AND customer_id=? AND visit_date=?",
        (user['id'], customer_id, visit_date)
    ).fetchone()

    if not visit or not visit['check_in_time']:
        db.close()
        return JSONResponse({"error": "לא נרשמה כניסה"})

    checkin_dt  = datetime.strptime(f"{visit_date} {visit['check_in_time']}", "%Y-%m-%d %H:%M:%S")
    checkout_dt = datetime.strptime(f"{visit_date} {now}", "%Y-%m-%d %H:%M:%S")
    if checkout_dt < checkin_dt:
        checkout_dt += timedelta(days=1)
    duration = max(0, int((checkout_dt - checkin_dt).total_seconds() / 60))

    # אם הביקור מושהה — סגור את ההשהייה אוטומטית לפני היציאה
    try:
        paused_at = visit['paused_at']
    except Exception:
        paused_at = None
    if paused_at:
        pause_start = datetime.strptime(f"{visit_date} {paused_at}", "%Y-%m-%d %H:%M:%S")
        extra_pause = max(1, int((checkout_dt - pause_start).total_seconds() / 60))
        try:
            cur_pause = visit['pause_duration_minutes'] or 0
        except Exception:
            cur_pause = 0
        db.execute(
            "UPDATE visits SET paused_at=NULL, pause_duration_minutes=? WHERE id=?",
            (cur_pause + extra_pause, visit['id'])
        )
        db.commit()

    db.execute(
        "UPDATE visits SET check_out_time=?, duration_minutes=?, notes=? WHERE id=?",
        (now, duration, notes, visit['id'])
    )

    customer = db.execute("SELECT name FROM customers WHERE id=?", (customer_id,)).fetchone()
    store_name = customer['name'] if customer else str(customer_id)

    hours = duration // 60
    mins = duration % 60
    duration_str = f"{hours} שעות {mins} דקות" if hours > 0 else f"{mins} דקות"

    db.execute(
        "INSERT INTO notifications (message, user_name, store_name, action, created_at) VALUES (?,?,?,?,?)",
        (f"{user['name']} יצא מ{store_name} — שהה {duration_str}", user['name'], store_name, 'checkout', now_il().isoformat())
    )
    db.commit()
    db.close()

    return JSONResponse({"status": "ok", "duration": duration, "duration_str": duration_str})

@app.post("/api/notes")
async def save_notes(request: Request):
    user = get_current_user(request)
    if not user:
        return JSONResponse({"error": "לא מחובר"}, status_code=401)

    data = await request.json()
    customer_id = data.get("customer_id")
    visit_date = data.get("visit_date")
    notes = data.get("notes", "")

    db = get_db()
    existing = db.execute(
        "SELECT id FROM visits WHERE user_id=? AND customer_id=? AND visit_date=?",
        (user['id'], customer_id, visit_date)
    ).fetchone()

    if existing:
        db.execute("UPDATE visits SET notes=? WHERE id=?", (notes, existing['id']))
    else:
        db.execute(
            "INSERT INTO visits (user_id, customer_id, visit_date, notes) VALUES (?,?,?,?)",
            (user['id'], customer_id, visit_date, notes)
        )
    db.commit()
    db.close()
    return JSONResponse({"status": "ok"})

# ─── MANAGER ─────────────────────────────────────────────────────────────────

@app.post("/api/comment")
async def add_comment(request: Request):
    user = get_current_user(request)
    if not user:
        return JSONResponse({"error": "לא מחובר"}, status_code=401)
    data = await request.json()
    customer_id = data.get("customer_id")
    visit_date = data.get("visit_date")
    message = data.get("message", "").strip()
    if not message:
        return JSONResponse({"error": "הודעה ריקה"})
    db = get_db()
    visit = db.execute(
        "SELECT id FROM visits WHERE customer_id=? AND visit_date=?",
        (customer_id, visit_date)
    ).fetchone()
    visit_id = visit['id'] if visit else None
    db.execute(
        "INSERT INTO visit_comments (visit_id, customer_id, visit_date, user_id, user_name, user_role, message, created_at) VALUES (?,?,?,?,?,?,?,?)",
        (visit_id, customer_id, visit_date, user['id'], user['name'], user['role'], message, now_il().isoformat())
    )
    db.commit()
    # Push notification to manager if agent sent message
    if user['role'] == 'agent':
        customer_row = db.execute("SELECT name FROM customers WHERE id=?", (customer_id,)).fetchone()
        store_label = customer_row['name'] if customer_row else f"חנות #{customer_id}"
        db.close()
        try:
            # URL — only numeric IDs to avoid encoding issues on iOS
            _push_url = f"/manager?open_chat={customer_id}&vdate={visit_date}"
            _send_push_to_managers(
                title=f"💬 הודעה מ{user['name']}",
                body=f"{store_label}: {message[:80]}",
                url=_push_url
            )
        except Exception:
            pass
    else:
        db.close()
    return JSONResponse({"status": "ok", "user_name": user['name'], "user_role": user['role'],
                         "message": message, "time": now_il().strftime("%H:%M")})

@app.get("/api/comments")
async def get_comments(request: Request, customer_id: int, visit_date: str):
    user = get_current_user(request)
    if not user:
        return JSONResponse({"error": "לא מחובר"}, status_code=401)
    db = get_db()
    comments = db.execute(
        "SELECT * FROM visit_comments WHERE customer_id=? AND visit_date=? ORDER BY created_at ASC",
        (customer_id, visit_date)
    ).fetchall()
    db.close()
    return JSONResponse({"comments": [dict(c) for c in comments]})


@app.get("/manager", response_class=HTMLResponse)
async def manager_dashboard(request: Request, date: str = None, agent_id: int = None):
    user = get_current_user(request)
    if not user or user['role'] != 'manager':
        return RedirectResponse("/")

    try:
        selected_date = datetime.strptime(date, "%Y-%m-%d").date() if date else now_il().date()
    except ValueError:
        selected_date = now_il().date()

    db = get_db()
    agents = db.execute("SELECT * FROM users WHERE role='agent' ORDER BY name").fetchall()

    # Visits for selected date
    visits_query = "SELECT v.*, u.name as agent_name, c.name as store_name, c.city, c.region FROM visits v JOIN users u ON v.user_id=u.id JOIN customers c ON v.customer_id=c.id WHERE v.visit_date=?"
    params = [str(selected_date)]
    if agent_id:
        visits_query += " AND v.user_id=?"
        params.append(agent_id)
    visits_query += " ORDER BY v.check_in_time DESC"

    visits = db.execute(visits_query, params).fetchall()

    # Notifications (last 50)
    notifications = db.execute(
        "SELECT * FROM notifications ORDER BY created_at DESC LIMIT 50"
    ).fetchall()

    # EOD reports — today only
    eod_reports = db.execute(
        """SELECT * FROM notifications
           WHERE action='eod_report' AND created_at LIKE ?
           ORDER BY created_at DESC""",
        (str(selected_date) + '%',)
    ).fetchall()

    unread = db.execute("SELECT COUNT(*) FROM notifications WHERE is_read=0").fetchone()[0]

    # Agent summary for today
    agent_summary = []
    for agent in agents:
        today_visits = db.execute(
            "SELECT COUNT(*) FROM visits WHERE user_id=? AND visit_date=? AND (check_out_time IS NOT NULL OR is_phone=1)",
            (agent['id'], str(selected_date))
        ).fetchone()[0]
        active_count = db.execute(
            "SELECT COUNT(*) FROM visits WHERE user_id=? AND visit_date=? AND check_in_time IS NOT NULL AND check_out_time IS NULL AND is_phone=0",
            (agent['id'], str(selected_date))
        ).fetchone()[0]
        current = db.execute(
            """SELECT v.*, c.name as store_name FROM visits v JOIN customers c ON v.customer_id=c.id
               WHERE v.user_id=? AND v.visit_date=? AND v.check_in_time IS NOT NULL AND v.check_out_time IS NULL""",
            (agent['id'], str(selected_date))
        ).fetchone()
        agent_regions = [r.strip() for r in (agent['regions'] or '').split(',') if r.strip()]
        if agent_regions:
            hebrew_day_sel = get_hebrew_day(selected_date)
            week_num_sel   = get_week_for_agent(selected_date, agent)
            week_col_sel   = f"week_{week_num_sel}"
            placeholders   = ','.join(['?' for _ in agent_regions])
            total_customers = db.execute(
                f"SELECT COUNT(*) FROM customers WHERE region IN ({placeholders}) AND assigned_visit_day=? AND {week_col_sel}=1",
                agent_regions + [hebrew_day_sel]
            ).fetchone()[0]
            if total_customers == 0:
                total_customers = db.execute(
                    f"SELECT COUNT(*) FROM customers WHERE region IN ({placeholders}) AND visit_day=?",
                    agent_regions + [hebrew_day_sel]
                ).fetchone()[0]
        else:
            total_customers = 0
        not_visited_count = max(0, total_customers - today_visits - active_count)
        agent_summary.append({
            "id": agent['id'],
            "name": agent['name'],
            "username": agent['username'],
            "regions": agent['regions'],
            "today_visits": today_visits,
            "total_customers": total_customers,
            "current_store": dict(current) if current else None,
            "done_count": today_visits,
            "active_count": active_count,
            "not_visited_count": not_visited_count,
        })

    db.close()

    prev_date = (selected_date - timedelta(days=1)).strftime("%Y-%m-%d")
    next_date = (selected_date + timedelta(days=1)).strftime("%Y-%m-%d")

    return templates.TemplateResponse("manager.html", {
        "request": request,
        "user": user,
        "agents": [dict(a) for a in agents],
        "agent_summary": agent_summary,
        "eod_reports": [dict(r) for r in eod_reports],
        "visits": [dict(v) for v in visits],
        "notifications": [dict(n) for n in notifications],
        "unread": unread,
        "selected_date": selected_date,
        "selected_date_str": str(selected_date),
        "prev_date": prev_date,
        "next_date": next_date,
        "selected_agent_id": agent_id,
        "today": now_il().date(),
        "manager_email": get_setting("manager_email", COMPANY_CONFIG.get("manager_email", "ranaz@matrix.co.il")),
        "cfg": COMPANY_CONFIG,
    })

@app.get("/api/agents-list")
async def agents_list(request: Request):
    """רשימת סוכנים — משמש לפתיחת צ'אט מנוטיפיקציה"""
    user = get_current_user(request)
    if not user or user['role'] != 'manager':
        return JSONResponse({"error": "אין הרשאה"}, status_code=403)
    db = get_db()
    rows = db.execute("SELECT id, name FROM users WHERE role='agent' ORDER BY name").fetchall()
    db.close()
    return JSONResponse([{"id": r['id'], "name": r['name']} for r in rows])

@app.post("/api/general-chat")
async def general_chat(request: Request):
    """צ'אט כללי בין סוכן למנהל (ללא קשר לחנות)"""
    user = get_current_user(request)
    if not user:
        return JSONResponse({"error": "לא מחובר"}, status_code=401)
    data = await request.json()
    message = data.get("message", "").strip()
    agent_id = data.get("agent_id")  # מנהל שולח לסוכן ספציפי
    if not message:
        return JSONResponse({"error": "הודעה ריקה"})
    # customer_id=0 = צ'אט כללי
    # visit_date = agent user id (כדי להפריד בין סוכנים)
    target_agent_id = agent_id if (user['role'] == 'manager' and agent_id) else user['id']
    db = get_db()
    db.execute(
        "INSERT INTO visit_comments (visit_id, customer_id, visit_date, user_id, user_name, user_role, message, created_at) VALUES (?,?,?,?,?,?,?,?)",
        (None, 0, str(target_agent_id), user['id'], user['name'], user['role'], message, now_il().isoformat())
    )
    db.commit()
    db.close()
    # Push notification למנהל כשסוכן שולח הודעה
    if user['role'] == 'agent':
        try:
            _send_push_to_managers(
                title=f"💬 הודעה מ{user['name']}",
                body=message[:100],
                url=f"/manager?open_mgc={user['id']}"
            )
        except Exception:
            pass
    return JSONResponse({"status": "ok"})

@app.delete("/api/general-chat/clear")
async def clear_general_chat(request: Request, agent_id: int = 0):
    """מחיקת כל הצ'אט הכללי — סוכן מוחק שלו, מנהל מוחק לפי agent_id"""
    user = get_current_user(request)
    if not user:
        return JSONResponse({"error": "לא מחובר"}, status_code=401)
    if user['role'] == 'manager' and agent_id:
        pass  # agent_id מגיע מה-query param
    else:
        agent_id = user['id']
    db = get_db()
    try:
        db.execute(
            "DELETE FROM visit_comments WHERE customer_id=0 AND visit_date=?",
            (str(agent_id),)
        )
        db.commit()
    except Exception:
        pass
    db.close()
    return JSONResponse({"status": "ok"})

@app.get("/api/my-customers")
async def get_my_customers(request: Request):
    """סוכן — רשימת כל הלקוחות שלו לפי אזורים"""
    user = get_current_user(request)
    if not user or user['role'] != 'agent':
        return JSONResponse({"error": "אין הרשאה"}, status_code=403)
    regions = [r.strip() for r in (user['regions'] or '').split(',') if r.strip()]
    if not regions:
        return JSONResponse({"customers": []})
    db = get_db()
    placeholders = ','.join(['?' for _ in regions])
    customers = db.execute(
        f"SELECT id, name, city, region, assigned_visit_day, visit_day FROM customers "
        f"WHERE region IN ({placeholders}) AND assigned_visit_day IS NOT NULL AND assigned_visit_day != '' "
        f"ORDER BY assigned_visit_day, name",
        regions
    ).fetchall()
    db.close()
    return JSONResponse({"customers": [dict(c) for c in customers]})

@app.get("/api/general-chat")
async def get_general_chat(request: Request, agent_id: int = None, mark_read: int = 0):
    """קבלת הצ'אט הכללי"""
    user = get_current_user(request)
    if not user:
        return JSONResponse({"error": "לא מחובר"}, status_code=401)
    target = agent_id if (user['role'] == 'manager' and agent_id) else user['id']
    db = get_db()
    comments = db.execute(
        "SELECT * FROM visit_comments WHERE customer_id=0 AND visit_date=? ORDER BY created_at ASC",
        (str(target),)
    ).fetchall()
    if mark_read:
        try:
            if user['role'] == 'agent':
                # סוכן פתח צ'אט → מסמן הודעות המנהל כנקראו
                db.execute(
                    "UPDATE visit_comments SET is_read=1 WHERE customer_id=0 AND visit_date=? AND user_role='manager'",
                    (str(user['id']),)
                )
            elif user['role'] == 'manager' and agent_id:
                # מנהל פתח צ'אט עם סוכן → מסמן הודעות הסוכן הזה כנקראו
                db.execute(
                    "UPDATE visit_comments SET is_read=1 WHERE customer_id=0 AND visit_date=? AND user_role='agent'",
                    (str(agent_id),)
                )
            db.commit()
        except Exception:
            pass
    db.close()
    return JSONResponse({"comments": [dict(c) for c in comments]})

@app.get("/api/manager-chat-unread-count")
async def manager_chat_unread_count(request: Request):
    """מנהל — ספירת צ'אטים כלליים מסוכנים שלא נקראו + אילו סוכנים"""
    user = get_current_user(request)
    if not user or user['role'] != 'manager':
        return JSONResponse({"count": 0, "agents_with_unread": []})
    db = get_db()
    try:
        count = db.execute(
            "SELECT COUNT(*) FROM visit_comments WHERE customer_id=0 AND user_role='agent' AND (is_read IS NULL OR is_read=0)"
        ).fetchone()[0]
        rows = db.execute(
            "SELECT DISTINCT user_id FROM visit_comments WHERE customer_id=0 AND user_role='agent' AND (is_read IS NULL OR is_read=0)"
        ).fetchall()
        agents_with_unread = [r[0] for r in rows]
    except Exception:
        count = 0
        agents_with_unread = []
    db.close()
    return JSONResponse({"count": count, "agents_with_unread": agents_with_unread})

@app.get("/api/agent-unread-count")
async def agent_unread_count(request: Request):
    """סוכן — ספירת הודעות מנהל שלא נקראו"""
    user = get_current_user(request)
    if not user or user['role'] != 'agent':
        return JSONResponse({"count": 0})
    db = get_db()
    try:
        count = db.execute(
            "SELECT COUNT(*) FROM visit_comments WHERE customer_id=0 AND visit_date=? AND user_role='manager' AND (is_read IS NULL OR is_read=0)",
            (str(user['id']),)
        ).fetchone()[0]
    except Exception:
        count = 0
    db.close()
    return JSONResponse({"count": count})

@app.get("/api/unread-chats")
async def unread_chats(request: Request):
    """מנהל — ספירת צ'אטים שלא נענו מהסוכנים"""
    user = get_current_user(request)
    if not user or user['role'] != 'manager':
        return JSONResponse({"count": 0, "items": []})
    db = get_db()
    # הודעות מסוכנים שלא קיבלו תגובת מנהל אחריהן
    today = str(now_il().date())
    # כל הודעות אחרונות מסוכנים (בצ'אט של חנויות + כללי) מהיום
    recent = db.execute(
        """SELECT vc.*, c.name as store_name
           FROM visit_comments vc
           LEFT JOIN customers c ON vc.customer_id = c.id
           WHERE vc.user_role='agent' AND vc.created_at >= ?
           ORDER BY vc.created_at DESC""",
        (today + 'T00:00:00',)
    ).fetchall()
    db.close()
    items = []
    for r in recent:
        items.append({
            "id": r['id'],
            "customer_id": r['customer_id'],
            "visit_date": r['visit_date'],
            "store_name": r['store_name'] or 'צ\'אט כללי',
            "agent_name": r['user_name'],
            "agent_id": r['user_id'],
            "message": r['message'],
            "time": r['created_at'][11:16]
        })
    return JSONResponse({"count": len(items), "items": items[:10]})

@app.post("/api/pause")
async def pause_visit(request: Request):
    """השהיית ביקור — שמירת שעת תחילת ההשהייה"""
    user = get_current_user(request)
    if not user:
        return JSONResponse({"error": "לא מחובר"}, status_code=401)
    data = await request.json()
    customer_id = data.get("customer_id")
    visit_date  = data.get("visit_date")
    now = now_il().strftime("%H:%M:%S")
    db = get_db()
    visit = db.execute(
        "SELECT * FROM visits WHERE user_id=? AND customer_id=? AND visit_date=?",
        (user['id'], customer_id, visit_date)
    ).fetchone()
    if not visit or not visit['check_in_time'] or visit['check_out_time']:
        db.close()
        return JSONResponse({"error": "אין ביקור פעיל"})
    try:
        already_paused = visit['paused_at']
    except Exception:
        already_paused = None
    if already_paused:
        db.close()
        return JSONResponse({"error": "הביקור כבר מושהה"})
    db.execute(
        "UPDATE visits SET paused_at=? WHERE user_id=? AND customer_id=? AND visit_date=?",
        (now, user['id'], customer_id, visit_date)
    )
    db.commit()
    db.close()
    return JSONResponse({"status": "ok", "time": now})

@app.post("/api/resume")
async def resume_visit(request: Request):
    """המשך ביקור לאחר השהייה — חישוב זמן ההשהייה"""
    user = get_current_user(request)
    if not user:
        return JSONResponse({"error": "לא מחובר"}, status_code=401)
    data = await request.json()
    customer_id = data.get("customer_id")
    visit_date  = data.get("visit_date")
    now = now_il()
    db = get_db()
    visit = db.execute(
        "SELECT * FROM visits WHERE user_id=? AND customer_id=? AND visit_date=?",
        (user['id'], customer_id, visit_date)
    ).fetchone()
    try:
        paused_at = visit['paused_at'] if visit else None
    except Exception:
        paused_at = None
    if not paused_at:
        db.close()
        return JSONResponse({"error": "אין השהייה פעילה"})
    pause_start = datetime.strptime(f"{visit_date} {paused_at}", "%Y-%m-%d %H:%M:%S")
    now_cmp     = now if now >= pause_start else now + timedelta(days=1)
    pause_minutes = max(1, int((now_cmp - pause_start).total_seconds() / 60))
    try:
        current_total = visit['pause_duration_minutes'] or 0
    except Exception:
        current_total = 0
    new_total = current_total + pause_minutes
    db.execute(
        "UPDATE visits SET paused_at=NULL, pause_duration_minutes=? WHERE user_id=? AND customer_id=? AND visit_date=?",
        (new_total, user['id'], customer_id, visit_date)
    )
    db.commit()
    db.close()
    return JSONResponse({"status": "ok", "pause_minutes": pause_minutes, "total_pause": new_total})

@app.post("/api/toggle_phone")
async def toggle_phone(request: Request):
    user = get_current_user(request)
    if not user:
        return JSONResponse({"error": "לא מחובר"}, status_code=401)
    data = await request.json()
    customer_id = data.get("customer_id")
    visit_date  = data.get("visit_date", str(now_il().date()))
    db = get_db()
    visit = db.execute(
        "SELECT * FROM visits WHERE user_id=? AND customer_id=? AND visit_date=?",
        (user['id'], customer_id, visit_date)
    ).fetchone()
    if visit:
        new_val = 0 if (visit['is_phone'] or 0) else 1
        if new_val == 1 and not visit['check_in_time']:
            # ביקור ללא כניסה (הערה בלבד) — סמן כשיחה טלפונית ואכלס זמנים
            now = now_il().strftime("%H:%M:%S")
            db.execute(
                "UPDATE visits SET is_phone=1, check_in_time=?, check_out_time=?, duration_minutes=0 WHERE id=?",
                (now, now, visit['id'])
            )
        elif new_val == 0 and visit['check_in_time'] and visit['check_in_time'] == visit['check_out_time']:
            # ביטול שיחה טלפונית שנוצרה ע"י הכפתור (check_in==check_out) — אפס את הזמנים
            db.execute(
                "UPDATE visits SET is_phone=0, check_in_time=NULL, check_out_time=NULL, duration_minutes=NULL WHERE id=?",
                (visit['id'],)
            )
        else:
            db.execute("UPDATE visits SET is_phone=? WHERE id=?", (new_val, visit['id']))
    else:
        now = now_il().strftime("%H:%M:%S")
        db.execute(
            "INSERT INTO visits (user_id, customer_id, visit_date, check_in_time, check_out_time, duration_minutes, is_phone) VALUES (?,?,?,?,?,?,?)",
            (user['id'], customer_id, visit_date, now, now, 0, 1)
        )
        new_val = 1
    db.commit()
    db.close()
    return JSONResponse({"status": "ok", "is_phone": new_val})

@app.post("/api/reset-visit")
async def reset_visit(request: Request):
    """איפוס ביקור בודד לחנות ביום מסוים"""
    user = get_current_user(request)
    if not user:
        return JSONResponse({"error": "לא מחובר"}, status_code=401)
    data = await request.json()
    customer_id = data.get("customer_id")
    visit_date  = data.get("visit_date")
    db = get_db()
    # מנהל יכול לאפס כל ביקור; סוכן רק שלו
    if user['role'] == 'manager':
        db.execute("DELETE FROM visits WHERE customer_id=? AND visit_date=?",
                   (customer_id, visit_date))
        db.execute("DELETE FROM visit_comments WHERE customer_id=? AND visit_date=?",
                   (customer_id, visit_date))
    else:
        db.execute("DELETE FROM visits WHERE customer_id=? AND visit_date=? AND user_id=?",
                   (customer_id, visit_date, user['id']))
        db.execute("DELETE FROM visit_comments WHERE customer_id=? AND visit_date=? AND user_id=?",
                   (customer_id, visit_date, user['id']))
    db.commit()
    db.close()
    return JSONResponse({"status": "ok"})

@app.post("/api/reset-day")
async def reset_day(request: Request):
    """איפוס כל ביקורי היום לסוכן"""
    user = get_current_user(request)
    if not user:
        return JSONResponse({"error": "לא מחובר"}, status_code=401)
    data       = await request.json()
    visit_date = data.get("visit_date")
    agent_id   = data.get("agent_id")          # מנהל יכול לאפס לכל סוכן
    if user['role'] == 'manager' and agent_id:
        target_id = agent_id
    else:
        target_id = user['id']
    db = get_db()
    # קודם שומרים את רשימת החנויות שהסוכן ביקר בהן
    customer_ids = [row[0] for row in db.execute(
        "SELECT DISTINCT customer_id FROM visits WHERE user_id=? AND visit_date=?",
        (target_id, visit_date)
    ).fetchall()]
    # מחיקת ביקורים
    db.execute("DELETE FROM visits WHERE user_id=? AND visit_date=?",
               (target_id, visit_date))
    # מחיקת כל הצ'אטים לאותן חנויות ביום זה (של כולם — סוכן + מנהל)
    for cid in customer_ids:
        db.execute("DELETE FROM visit_comments WHERE customer_id=? AND visit_date=?",
                   (cid, visit_date))
    db.commit()
    db.close()
    return JSONResponse({"status": "ok"})

@app.get("/api/customers/search")
async def search_customers_agent(request: Request, q: str = ""):
    user = get_current_user(request)
    if not user:
        raise HTTPException(status_code=401)
    if len(q) < 2:
        return JSONResponse([])
    db = get_db()
    try:
        # מנהל — כל הלקוחות; סוכן — רק האזורים שלו
        if user['role'] == 'manager':
            rows = db.execute(
                "SELECT id, name, city, region, visit_day, assigned_visit_day FROM customers WHERE name LIKE ? ORDER BY name LIMIT 12",
                (f'%{q}%',)
            ).fetchall()
        else:
            regions = [r.strip() for r in user['regions'].split(',') if r.strip()]
            if not regions:
                return JSONResponse([])
            placeholders = ','.join(['?' for _ in regions])
            sql_search = (
                "SELECT id, name, city, region, visit_day, assigned_visit_day FROM customers WHERE name LIKE ? AND region IN (" + placeholders + ") ORDER BY name LIMIT 12"
            )
            rows = db.execute(
                sql_search,
                [f'%{q}%'] + regions
            ).fetchall()

        # מיפוי אזור → שם סוכן
        agents = db.execute("SELECT name, regions FROM users WHERE role='agent'").fetchall()
        region_to_agent = {}
        for a in agents:
            for r in (a['regions'] or '').split(','):
                region_to_agent[r.strip()] = a['name']

        return JSONResponse([{
            "id": r["id"], "name": r["name"],
            "city": r["city"] or "", "region": r["region"] or "",
            "agent": region_to_agent.get(r["region"] or "", "—"),
            "visit_day": r["assigned_visit_day"] or r["visit_day"] or ""
        } for r in rows])
    finally:
        db.close()


@app.get("/api/customer-history/{customer_id}")
async def customer_history(request: Request, customer_id: int):
    user = get_current_user(request)
    if not user:
        raise HTTPException(status_code=401)
    db = get_db()
    try:
        customer = db.execute("SELECT name FROM customers WHERE id=?", (customer_id,)).fetchone()
        visits = db.execute("""
            SELECT v.id, v.visit_date, v.check_in_time, v.check_out_time,
                   v.duration_minutes, v.notes, u.name as agent_name
            FROM visits v
            JOIN users u ON v.user_id = u.id
            WHERE v.customer_id = ?
            ORDER BY v.visit_date DESC, v.check_in_time DESC
        """, (customer_id,)).fetchall()

        result = []
        for v in visits:
            comments = db.execute("""
                SELECT user_name, user_role, message, created_at
                FROM visit_comments
                WHERE customer_id=? AND visit_date=?
                ORDER BY created_at ASC
            """, (customer_id, v['visit_date'])).fetchall()
            result.append({
                "date": v['visit_date'],
                "agent_name": v['agent_name'],
                "check_in":  (v['check_in_time']  or '')[:5],
                "check_out": (v['check_out_time'] or '')[:5],
                "duration":  v['duration_minutes'] or 0,
                "notes":     v['notes'] or '',
                "comments":  [{"user": c['user_name'], "role": c['user_role'],
                               "msg": c['message'], "time": c['created_at'][11:16]} for c in comments]
            })
        return JSONResponse({"customer_name": customer['name'] if customer else '', "visits": result})
    finally:
        db.close()

@app.get("/api/notifications")
async def get_notifications(request: Request, agent: str = ""):
    user = get_current_user(request)
    if not user or user['role'] != 'manager':
        return JSONResponse({"error": "אין הרשאה"}, status_code=403)

    db = get_db()
    if agent:
        notifs = db.execute(
            "SELECT * FROM notifications WHERE user_name=? ORDER BY created_at DESC LIMIT 50",
            (agent,)
        ).fetchall()
    else:
        notifs = db.execute(
            "SELECT * FROM notifications ORDER BY created_at DESC LIMIT 50"
        ).fetchall()
    unread = db.execute("SELECT COUNT(*) FROM notifications WHERE is_read=0").fetchone()[0]
    db.execute("UPDATE notifications SET is_read=1")
    db.commit()
    db.close()
    return JSONResponse({"notifications": [dict(n) for n in notifs], "unread": unread})

@app.get("/api/unread-count")
async def unread_count(request: Request):
    """מספר הודעות סוכנים שלא נקראו — לצביעת כפתור ההתראות"""
    user = get_current_user(request)
    if not user or user['role'] != 'manager':
        return JSONResponse({"count": 0})
    db = get_db()
    err = None
    try:
        count = db.execute(
            "SELECT COUNT(*) FROM visit_comments WHERE user_role='agent' AND customer_id > 0 AND (is_read IS NULL OR is_read=0)"
        ).fetchone()[0]
    except Exception as e:
        count = 0
        err = str(e)
    db.close()
    return JSONResponse({"count": count, "error": err})

@app.get("/api/manager/beep")
async def get_beep(request: Request):
    user = get_current_user(request)
    if not user or user['role'] != 'manager':
        return JSONResponse({"error": "unauthorized"}, status_code=403)
    db = get_db()
    row = db.execute("SELECT beep_sound FROM users WHERE id=?", (user['id'],)).fetchone()
    db.close()
    return JSONResponse({"beep": row['beep_sound'] if row and row['beep_sound'] else 4})

@app.post("/api/manager/beep")
async def set_beep(request: Request):
    user = get_current_user(request)
    if not user or user['role'] != 'manager':
        return JSONResponse({"error": "unauthorized"}, status_code=403)
    data = await request.json()
    beep = int(data.get("beep", 4))
    db = get_db()
    db.execute("UPDATE users SET beep_sound=? WHERE id=?", (beep, user['id']))
    db.commit()
    db.close()
    return JSONResponse({"ok": True})

@app.get("/api/manager-summary")
async def manager_summary_api(request: Request, date: str = ""):
    user = get_current_user(request)
    if not user or user['role'] != 'manager':
        return JSONResponse({"error": "unauthorized"}, status_code=403)
    db = get_db()
    try:
        selected_date = datetime.strptime(date, "%Y-%m-%d").date() if date else now_il().date()
    except Exception:
        selected_date = now_il().date()
    agents = db.execute("SELECT * FROM users WHERE role='agent'").fetchall()
    summary = []
    for agent in agents:
        today_visits = db.execute(
            "SELECT COUNT(*) FROM visits WHERE user_id=? AND visit_date=? AND (check_out_time IS NOT NULL OR is_phone=1)",
            (agent['id'], str(selected_date))
        ).fetchone()[0]
        active_count = db.execute(
            "SELECT COUNT(*) FROM visits WHERE user_id=? AND visit_date=? AND check_in_time IS NOT NULL AND check_out_time IS NULL AND is_phone=0",
            (agent['id'], str(selected_date))
        ).fetchone()[0]
        current = db.execute(
            """SELECT c.name as store_name FROM visits v JOIN customers c ON v.customer_id=c.id
               WHERE v.user_id=? AND v.visit_date=? AND v.check_in_time IS NOT NULL AND v.check_out_time IS NULL""",
            (agent['id'], str(selected_date))
        ).fetchone()
        agent_regions = [r.strip() for r in (agent['regions'] or '').split(',') if r.strip()]
        if agent_regions:
            hebrew_day_sel = get_hebrew_day(selected_date)
            week_num_sel   = get_week_for_agent(selected_date, agent)
            week_col_sel   = f"week_{week_num_sel}"
            placeholders   = ','.join(['?' for _ in agent_regions])
            total_customers = db.execute(
                f"SELECT COUNT(*) FROM customers WHERE region IN ({placeholders}) AND assigned_visit_day=? AND {week_col_sel}=1",
                agent_regions + [hebrew_day_sel]
            ).fetchone()[0]
            if total_customers == 0:
                total_customers = db.execute(
                    f"SELECT COUNT(*) FROM customers WHERE region IN ({placeholders}) AND visit_day=?",
                    agent_regions + [hebrew_day_sel]
                ).fetchone()[0]
        else:
            total_customers = 0
        not_visited = max(0, total_customers - today_visits - active_count)
        summary.append({
            "id": agent['id'],
            "today_visits": today_visits,
            "total_customers": total_customers,
            "current_store": dict(current)['store_name'] if current else None,
            "done_count": today_visits,
            "active_count": active_count,
            "not_visited_count": not_visited,
        })
    db.close()
    return JSONResponse({"agents": summary})

@app.get("/api/agent-messages")
async def agent_messages(request: Request, agent: str = "", mark_read: str = "0"):
    """הודעות צ'אט של סוכנים — לטאב ההתראות"""
    user = get_current_user(request)
    if not user or user['role'] != 'manager':
        return JSONResponse({"error": "אין הרשאה"}, status_code=403)
    db = get_db()
    if agent:
        msgs = db.execute(
            """SELECT vc.*, c.name as store_name
               FROM visit_comments vc
               LEFT JOIN customers c ON vc.customer_id = c.id
               WHERE vc.user_role='agent' AND vc.customer_id > 0 AND vc.user_name=?
               ORDER BY vc.created_at DESC LIMIT 60""",
            (agent,)
        ).fetchall()
    else:
        msgs = db.execute(
            """SELECT vc.*, c.name as store_name
               FROM visit_comments vc
               LEFT JOIN customers c ON vc.customer_id = c.id
               WHERE vc.user_role='agent' AND vc.customer_id > 0
               ORDER BY vc.created_at DESC LIMIT 60"""
        ).fetchall()
    # סמן כנקרא רק כשהמנהל פתח את הטאב בפועל
    if mark_read == "1":
        # מסמן רק הודעות חנות כנקראו — לא צ'אט כללי (customer_id=0)
        db.execute("UPDATE visit_comments SET is_read=1 WHERE user_role='agent' AND customer_id > 0")
        db.commit()
    db.close()
    return JSONResponse({"messages": [dict(m) for m in msgs]})

@app.get("/api/customer-name/{customer_id}")
async def get_customer_name(customer_id: int, request: Request):
    """שם חנות לפי ID — לשימוש ב-deep link מ-push notification"""
    user = get_current_user(request)
    if not user:
        return JSONResponse({"error": "לא מחובר"}, status_code=401)
    db = get_db()
    row = db.execute("SELECT name FROM customers WHERE id=?", (customer_id,)).fetchone()
    db.close()
    return JSONResponse({"name": row['name'] if row else f"חנות #{customer_id}"})

@app.post("/api/mark-chat-read")
async def mark_chat_read(request: Request):
    """מנהל — סמן הודעות של חנות+תאריך ספציפיים כנקראו"""
    user = get_current_user(request)
    if not user or user['role'] != 'manager':
        return JSONResponse({"error": "אין הרשאה"}, status_code=403)
    data = await request.json()
    customer_id = data.get("customer_id")
    visit_date  = data.get("visit_date")
    db = get_db()
    try:
        db.execute(
            "UPDATE visit_comments SET is_read=1 WHERE customer_id=? AND visit_date=? AND user_role='agent'",
            (customer_id, visit_date)
        )
        db.commit()
    except Exception:
        pass
    db.close()
    return JSONResponse({"status": "ok"})

@app.delete("/api/agent-messages/clear-all")
async def delete_all_agent_messages(request: Request):
    """מחיקת כל הודעות הסוכנים — מנהל בלבד"""
    user = get_current_user(request)
    if not user or user['role'] != 'manager':
        return JSONResponse({"error": "אין הרשאה"}, status_code=403)
    db = get_db()
    db.execute("DELETE FROM visit_comments WHERE customer_id > 0")
    db.commit()
    db.close()
    return JSONResponse({"status": "ok"})

@app.delete("/api/agent-messages/{message_id}")
async def delete_agent_message(message_id: int, request: Request):
    """מחיקת הודעת סוכן — מנהל בלבד"""
    user = get_current_user(request)
    if not user or user['role'] != 'manager':
        return JSONResponse({"error": "אין הרשאה"}, status_code=403)
    db = get_db()
    db.execute("DELETE FROM visit_comments WHERE id=?", (message_id,))
    db.commit()
    db.close()
    return JSONResponse({"status": "ok"})

# ─── EXCEL REPORTS ───────────────────────────────────────────────────────────

def _write_sheet(ws, title: str, subtitle: str, col_widths: list, headers: list, rows: list, summary: list):
    import re as _re
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    NAVY  = "00327d"; LIGHT = "EEF1F8"; ALT = "F0F5FF"; GOLD = "FFF9C4"; WHITE = "FFFFFF"
    n = len(headers)
    def T():
        s = Side(style="thin", color="B8C5E0")
        return Border(left=s, right=s, top=s, bottom=s)
    ws.sheet_view.rightToLeft = True
    last_col = chr(64 + n)
    ws.merge_cells(f"A1:{last_col}1")
    c = ws["A1"]; c.value = f"{COMPANY_CONFIG.get('company_name', 'SmartConnect')}  |  {title}"
    c.font = Font(bold=True, size=14, color=WHITE, name="Arial")
    c.fill = PatternFill("solid", start_color=NAVY)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30
    ws.merge_cells(f"A2:{last_col}2")
    c = ws["A2"]; c.value = subtitle
    c.font = Font(size=11, color=NAVY, name="Arial")
    c.fill = PatternFill("solid", start_color=LIGHT)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 22
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=3, column=col, value=h)
        c.font = Font(bold=True, color=WHITE, size=11, name="Arial")
        c.fill = PatternFill("solid", start_color=NAVY)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = T()
    ws.row_dimensions[3].height = 22
    for i, row in enumerate(rows):
        r = i + 4; bg = WHITE if i % 2 == 0 else ALT
        for col, val in enumerate(row, 1):
            c = ws.cell(row=r, column=col, value=val)
            c.font = Font(name="Arial", size=10)
            c.fill = PatternFill("solid", start_color=bg)
            c.border = T()
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=(col == n))
    for j, (lbl, val) in enumerate(summary):
        r = len(rows) + 4 + j
        ws.merge_cells(f"A{r}:{chr(64+n-1)}{r}")
        for col in range(1, n + 1):
            c = ws.cell(row=r, column=col)
            c.fill = PatternFill("solid", start_color=GOLD); c.border = T()
            c.font = Font(bold=True, size=11, color=NAVY, name="Arial")
            c.alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row=r, column=1).value = lbl; ws.cell(row=r, column=n).value = val
    for i, w in enumerate(col_widths):
        ws.column_dimensions[chr(65 + i)].width = w
    ws.freeze_panes = "A4"


def _build_excel(title: str, subtitle: str, col_widths: list, headers: list, rows: list, summary: list) -> io.BytesIO:
    import re as _re
    from openpyxl import Workbook
    wb = Workbook(); ws = wb.active
    ws.title = _re.sub(r'[\[\]:*?/\\]', '_', title)[:31]
    _write_sheet(ws, title, subtitle, col_widths, headers, rows, summary)
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return buf


def _build_excel_multi(sheets: list) -> io.BytesIO:
    """sheets: list of dicts {sheet_name, title, subtitle, col_widths, headers, rows, summary}"""
    import re as _re
    from openpyxl import Workbook
    wb = Workbook(); wb.remove(wb.active)
    for s in sheets:
        safe = _re.sub(r'[\[\]:*?/\\]', '_', s['sheet_name'])[:31] or 'גיליון'
        ws = wb.create_sheet(title=safe)
        _write_sheet(ws, s['title'], s['subtitle'], s['col_widths'], s['headers'], s['rows'], s['summary'])
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return buf


def _agent_visit_rows(db, agent_id: int, from_date: str, to_date: str):
    rows_db = db.execute(
        """SELECT v.visit_date, COALESCE(c.name,'—') as store_name,
                  COALESCE(c.city,'—') as city, COALESCE(c.region,'—') as region,
                  v.check_in_time, v.check_out_time, v.duration_minutes, v.notes, v.is_phone
           FROM visits v LEFT JOIN customers c ON v.customer_id=c.id
           WHERE v.user_id=? AND v.visit_date BETWEEN ? AND ?
             AND (v.check_in_time IS NOT NULL OR v.is_phone=1 OR v.notes IS NOT NULL)
           ORDER BY v.visit_date DESC, v.check_in_time""",
        (agent_id, from_date, to_date)
    ).fetchall()
    rows = []; total_min = 0
    for v in rows_db:
        dur = max(0, v['duration_minutes'] or 0); total_min += dur
        rows.append([
            v['visit_date'], v['store_name'], v['city'], v['region'],
            (v['check_in_time'] or '')[:5], (v['check_out_time'] or '')[:5],
            dur if v['check_out_time'] else '—',
            'טלפוני' if v['is_phone'] else ('הערה' if not v['check_in_time'] else 'פרונטלי'),
            v['notes'] or '',
        ])
    return rows, total_min


def _customer_visit_rows(db, customer_id: int, from_date: str, to_date: str):
    rows_db = db.execute(
        """SELECT v.visit_date, COALESCE(u.name,'—') as agent_name,
                  v.check_in_time, v.check_out_time, v.duration_minutes, v.notes, v.is_phone
           FROM visits v LEFT JOIN users u ON v.user_id=u.id
           WHERE v.customer_id=? AND v.visit_date BETWEEN ? AND ?
             AND (v.check_in_time IS NOT NULL OR v.is_phone=1 OR v.notes IS NOT NULL)
           ORDER BY v.visit_date DESC""",
        (customer_id, from_date, to_date)
    ).fetchall()
    rows = []; total_min = 0
    for v in rows_db:
        dur = v['duration_minutes'] or 0; total_min += dur
        rows.append([
            v['visit_date'], v['agent_name'],
            (v['check_in_time'] or '')[:5], (v['check_out_time'] or '')[:5],
            dur if dur else '',
            'טלפוני' if v['is_phone'] else ('הערה' if not v['check_in_time'] else 'פרונטלי'),
            v['notes'] or '',
        ])
    return rows, total_min


@app.get("/api/search-customers")
async def search_customers_manager(request: Request, q: str = ""):
    user = get_current_user(request)
    if not user or user['role'] != 'manager':
        return JSONResponse({"results": []})
    if len(q) < 1:
        return JSONResponse({"results": []})
    db = get_db()
    rows = db.execute(
        "SELECT id, name, city, card_code FROM customers WHERE name LIKE ? OR card_code LIKE ? LIMIT 20",
        (f"%{q}%", f"%{q}%")
    ).fetchall()
    db.close()
    return JSONResponse({"results": [dict(r) for r in rows]})


@app.get("/api/debug/visits")
async def debug_visits(request: Request, agent_id: int = 0, from_date: str = "", to_date: str = ""):
    user = get_current_user(request)
    if not user or user['role'] != 'manager':
        return JSONResponse({"error": "אין גישה"}, status_code=403)
    db = get_db()
    rows = db.execute(
        """SELECT v.id, v.visit_date, v.check_in_time, v.check_out_time, v.is_phone, v.notes,
                  c.name as store_name
           FROM visits v LEFT JOIN customers c ON v.customer_id=c.id
           WHERE v.user_id=? AND v.visit_date BETWEEN ? AND ?
           ORDER BY v.visit_date DESC""",
        (agent_id, from_date, to_date)
    ).fetchall()
    db.close()
    return JSONResponse({
        "count": len(rows),
        "agent_id": agent_id,
        "from_date": from_date,
        "to_date": to_date,
        "visits": [{"id": r["id"], "date": r["visit_date"], "store": r["store_name"],
                    "check_in": r["check_in_time"], "check_out": r["check_out_time"],
                    "is_phone": r["is_phone"], "notes": r["notes"]} for r in rows]
    })


@app.get("/manager/report/agent")
async def report_agent(request: Request, agent_id: int = 0, from_date: str = "", to_date: str = ""):
    import traceback
    user = get_current_user(request)
    if not user or user['role'] != 'manager':
        return RedirectResponse("/")

    try:
        db = get_db()
        agent = db.execute("SELECT * FROM users WHERE id=?", (agent_id,)).fetchone()
        if not agent:
            db.close()
            return JSONResponse({"error": "סוכן לא נמצא"}, status_code=404)

        rows_db = db.execute(
            """SELECT v.visit_date, COALESCE(c.name,'—') as store_name, COALESCE(c.city,'—') as city, COALESCE(c.region,'—') as region,
                      v.check_in_time, v.check_out_time, v.duration_minutes, v.notes, v.is_phone
               FROM visits v LEFT JOIN customers c ON v.customer_id=c.id
               WHERE v.user_id=? AND v.visit_date BETWEEN ? AND ?
                 AND (v.check_in_time IS NOT NULL OR v.is_phone=1 OR v.notes IS NOT NULL)
               ORDER BY v.visit_date DESC, v.check_in_time""",
            (agent_id, from_date, to_date)
        ).fetchall()
        db.close()

        rows = []
        total_min = 0
        for v in rows_db:
            dur = max(0, v['duration_minutes'] or 0)
            total_min += dur
            rows.append([
                v['visit_date'],
                v['store_name'],
                v['city'],
                v['region'],
                (v['check_in_time'] or '')[:5],
                (v['check_out_time'] or '')[:5],
                dur if v['check_out_time'] else '—',
                'טלפוני' if v['is_phone'] else ('הערה' if not v['check_in_time'] else 'פרונטלי'),
                v['notes'] or '',
            ])

        h = total_min // 60
        m = total_min % 60
        summary = [
            ("סה\"כ ביקורים", len(rows)),
            ("סה\"כ זמן שהייה", f"{h}:{m:02d} שעות"),
        ]
        headers    = ["תאריך", "שם חנות", "עיר", "אזור", "כניסה", "יציאה", "משך (דק')", "סוג", "הערות"]
        col_widths = [14,      28,        14,    14,    10,      10,      11,           10,    30]

        buf = _build_excel(
            title=f"דוח ביקורים - {agent['name']}",
            subtitle=f"תקופה: {from_date}  עד  {to_date}",
            col_widths=col_widths,
            headers=headers,
            rows=rows,
            summary=summary,
        )
        fname = f"agent_{agent['name']}_{from_date}_{to_date}.xlsx"
        fname_encoded = quote(fname, safe='')
        return StreamingResponse(
            iter([buf.getvalue()]),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename*=UTF-8''{fname_encoded}"}
        )
    except Exception:
        err = traceback.format_exc()
        print(f"[REPORT AGENT ERROR]\n{err}")
        return JSONResponse({"error": "שגיאה בהפקת הדוח", "detail": err}, status_code=500)


@app.get("/manager/report/agents")
async def report_agents_multi(request: Request, agent_ids: str = "", from_date: str = "", to_date: str = ""):
    import traceback
    user = get_current_user(request)
    if not user or user['role'] != 'manager':
        return RedirectResponse("/")
    try:
        ids = [int(x.strip()) for x in agent_ids.split(',') if x.strip().isdigit()]
        if not ids:
            return JSONResponse({"error": "לא נבחרו סוכנים"}, status_code=400)
        db = get_db()
        headers    = ["תאריך", "שם חנות", "עיר", "אזור", "כניסה", "יציאה", "משך (דק')", "סוג", "הערות"]
        col_widths = [14, 28, 14, 14, 10, 10, 11, 10, 30]
        sheets = []
        for aid in ids:
            agent = db.execute("SELECT * FROM users WHERE id=?", (aid,)).fetchone()
            if not agent: continue
            rows, total_min = _agent_visit_rows(db, aid, from_date, to_date)
            h, m = total_min // 60, total_min % 60
            sheets.append({
                'sheet_name': agent['name'],
                'title': f"דוח ביקורים - {agent['name']}",
                'subtitle': f"תקופה: {from_date}  עד  {to_date}",
                'col_widths': col_widths, 'headers': headers, 'rows': rows,
                'summary': [("סה\"כ ביקורים", len(rows)), ("סה\"כ זמן שהייה", f"{h}:{m:02d} שעות")],
            })
        db.close()
        if not sheets:
            return JSONResponse({"error": "לא נמצאו סוכנים"}, status_code=404)
        buf = _build_excel_multi(sheets)
        fname = quote(f"agents_report_{from_date}_{to_date}.xlsx", safe='')
        return StreamingResponse(iter([buf.getvalue()]),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename*=UTF-8''{fname}"})
    except Exception:
        err = traceback.format_exc()
        print(f"[REPORT AGENTS ERROR]\n{err}")
        return JSONResponse({"error": "שגיאה בהפקת הדוח", "detail": err}, status_code=500)


@app.get("/manager/report/customers")
async def report_customers_multi(request: Request, customer_ids: str = "", from_date: str = "", to_date: str = ""):
    import traceback
    user = get_current_user(request)
    if not user or user['role'] != 'manager':
        return RedirectResponse("/")
    try:
        ids = [int(x.strip()) for x in customer_ids.split(',') if x.strip().isdigit()]
        if not ids:
            return JSONResponse({"error": "לא נבחרו לקוחות"}, status_code=400)
        db = get_db()
        headers    = ["תאריך", "סוכן", "כניסה", "יציאה", "משך (דק')", "סוג", "הערות"]
        col_widths = [14, 18, 10, 10, 11, 10, 35]
        sheets = []
        for cid in ids:
            customer = db.execute("SELECT * FROM customers WHERE id=?", (cid,)).fetchone()
            if not customer: continue
            rows, total_min = _customer_visit_rows(db, cid, from_date, to_date)
            h, m = total_min // 60, total_min % 60
            cname = customer['name']; ccity = customer['city'] or ''
            sheets.append({
                'sheet_name': f"{cname[:20]}{' '+ccity if ccity else ''}"[:31],
                'title': f"דוח ביקורים - {cname}",
                'subtitle': f"תקופה: {from_date}  עד  {to_date}  |  {ccity}  |  {customer['region'] or ''}",
                'col_widths': col_widths, 'headers': headers, 'rows': rows,
                'summary': [("סה\"כ ביקורים", len(rows)), ("סה\"כ זמן שהייה", f"{h}:{m:02d} שעות")],
            })
        db.close()
        if not sheets:
            return JSONResponse({"error": "לא נמצאו לקוחות"}, status_code=404)
        buf = _build_excel_multi(sheets)
        fname = quote(f"customers_report_{from_date}_{to_date}.xlsx", safe='')
        return StreamingResponse(iter([buf.getvalue()]),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename*=UTF-8''{fname}"})
    except Exception:
        err = traceback.format_exc()
        print(f"[REPORT CUSTOMERS ERROR]\n{err}")
        return JSONResponse({"error": "שגיאה בהפקת הדוח", "detail": err}, status_code=500)


@app.get("/manager/report/customer")
async def report_customer(request: Request, customer_id: int = 0, from_date: str = "", to_date: str = ""):
    import traceback
    user = get_current_user(request)
    if not user or user['role'] != 'manager':
        return RedirectResponse("/")

    try:
        db = get_db()
        customer = db.execute("SELECT * FROM customers WHERE id=?", (customer_id,)).fetchone()
        if not customer:
            db.close()
            return JSONResponse({"error": "לקוח לא נמצא"}, status_code=404)

        rows_db = db.execute(
            """SELECT v.visit_date, u.name as agent_name,
                      v.check_in_time, v.check_out_time, v.duration_minutes, v.notes, v.is_phone
               FROM visits v JOIN users u ON v.user_id=u.id
               WHERE v.customer_id=? AND v.visit_date BETWEEN ? AND ?
                 AND v.check_in_time IS NOT NULL
               ORDER BY v.visit_date DESC""",
            (customer_id, from_date, to_date)
        ).fetchall()
        db.close()

        rows = []
        total_min = 0
        for v in rows_db:
            dur = v['duration_minutes'] or 0
            total_min += dur
            rows.append([
                v['visit_date'],
                v['agent_name'],
                (v['check_in_time'] or '')[:5],
                (v['check_out_time'] or '')[:5],
                dur if dur else '',
                'טלפוני' if v['is_phone'] else 'פרונטלי',
                v['notes'] or '',
            ])

        h = total_min // 60
        m = total_min % 60
        summary = [
            ("סה\"כ ביקורים", len(rows)),
            ("סה\"כ זמן שהייה", f"{h}:{m:02d} שעות"),
        ]
        headers    = ["תאריך", "סוכן", "כניסה", "יציאה", "משך (דק')", "סוג", "הערות"]
        col_widths = [14,      18,     10,      10,      11,           10,    35]

        buf = _build_excel(
            title=f"דוח ביקורים - {customer['name']}",
            subtitle=f"תקופה: {from_date}  עד  {to_date}  |  {customer['city'] or ''}  |  {customer['region'] or ''}",
            col_widths=col_widths,
            headers=headers,
            rows=rows,
            summary=summary,
        )
        fname = f"customer_{customer['name']}_{from_date}_{to_date}.xlsx"
        fname_encoded = quote(fname, safe='')
        return StreamingResponse(
            iter([buf.getvalue()]),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename*=UTF-8''{fname_encoded}"}
        )
    except Exception:
        err = traceback.format_exc()
        print(f"[REPORT CUSTOMER ERROR]\n{err}")
        return JSONResponse({"error": "שגיאה בהפקת הדוח", "detail": err}, status_code=500)


@app.get("/manager/report/schedule")
async def report_schedule(request: Request, month: str = ""):
    import traceback, calendar
    from datetime import date as _date
    user = get_current_user(request)
    if not user or user['role'] != 'manager':
        return RedirectResponse("/")
    try:
        if not month:
            month = now_il().date().strftime("%Y-%m")
        year, mon = map(int, month.split('-'))
        HEB_MONTHS = {1:'ינואר',2:'פברואר',3:'מרץ',4:'אפריל',5:'מאי',6:'יוני',
                      7:'יולי',8:'אוגוסט',9:'ספטמבר',10:'אוקטובר',11:'נובמבר',12:'דצמבר'}
        HEB_DAY   = {'א':'ראשון','ב':'שני','ג':'שלישי','ד':'רביעי','ה':'חמישי','ו':'שישי','ש':'שבת'}
        PY_TO_HEB = {0:'ב',1:'ג',2:'ד',3:'ה',4:'ו',5:'ש',6:'א'}
        _, days_in = calendar.monthrange(year, mon)
        all_dates  = [_date(year, mon, d) for d in range(1, days_in + 1)]
        month_disp = f"{HEB_MONTHS[mon]} {year}"
        db = get_db()
        try:
            agents = db.execute(
                "SELECT id, name, regions FROM users WHERE role='agent' ORDER BY name"
            ).fetchall()
            sheets = []
            summary_rows = []
            for agent in agents:
                regions = [r.strip() for r in (agent['regions'] or '').split(',') if r.strip()]
                if not regions:
                    continue
                ph = ','.join(['?'] * len(regions))
                customers = db.execute(
                    f"SELECT name, city, region, visit_day, week_1, week_2, week_3, week_4 "
                    f"FROM customers WHERE region IN ({ph}) ORDER BY visit_day, name",
                    regions
                ).fetchall()
                rows = []
                for d in all_dates:
                    heb = PY_TO_HEB[d.weekday()]
                    wk  = min((d.day - 1) // 7 + 1, 4)
                    for c in customers:
                        if (c['visit_day'] or '').strip() == heb and c[f'week_{wk}']:
                            rows.append([
                                d.strftime("%d/%m/%Y"),
                                HEB_DAY.get(heb, heb),
                                f'שבוע {wk}',
                                c['name'],
                                c['city'] or '',
                                c['region'] or '',
                            ])
                work_days = len({r[0] for r in rows})
                summary_rows.append([agent['name'], len(rows), len(customers), work_days])
                sheets.append({
                    'sheet_name': agent['name'],
                    'title': f"סידור עבודה — {agent['name']}",
                    'subtitle': f"חודש: {month_disp}  |  {len(rows)} ביקורים מתוכננים  |  {work_days} ימי עבודה",
                    'col_widths': [14, 12, 10, 30, 14, 14],
                    'headers': ['תאריך', 'יום', 'שבוע', 'שם לקוח', 'עיר', 'אזור'],
                    'rows': rows,
                    'summary': [(f'סה"כ ביקורים מתוכננים:', len(rows))],
                })
            sheets.insert(0, {
                'sheet_name': 'סיכום',
                'title': 'סיכום סידור עבודה חודשי',
                'subtitle': f'חודש: {month_disp}',
                'col_widths': [22, 18, 18, 18],
                'headers': ['סוכן', 'ביקורים מתוכננים', 'לקוחות באזור', 'ימי עבודה'],
                'rows': summary_rows,
                'summary': [('סה"כ ביקורים בחברה:', sum(r[1] for r in summary_rows))],
            })
            buf = _build_excel_multi(sheets)
            fname_encoded = quote(f"schedule_{month}.xlsx", safe='')
            return StreamingResponse(
                iter([buf.getvalue()]),
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={"Content-Disposition": f"attachment; filename*=UTF-8''{fname_encoded}"}
            )
        finally:
            db.close()
    except Exception:
        err = traceback.format_exc()
        print(f"[REPORT SCHEDULE ERROR]\n{err}")
        return JSONResponse({"error": "שגיאה בהפקת הדוח", "detail": err}, status_code=500)


@app.get("/api/test-email-config")
async def test_email_config(request: Request):
    """בדיקת הגדרות מייל — מחזיר סטטוס env vars + חיבור SMTP"""
    import traceback, smtplib
    user = get_current_user(request)
    if not user or user['role'] != 'manager':
        return JSONResponse({"error": "אין הרשאה"}, status_code=403)
    mail_from = os.environ.get("MAIL_FROM", "")
    mail_pass = os.environ.get("MAIL_PASSWORD", "")
    result = {
        "MAIL_FROM_set": bool(mail_from),
        "MAIL_FROM_value": mail_from[:4] + "***" if mail_from else "(ריק)",
        "MAIL_PASSWORD_set": bool(mail_pass),
    }
    if not mail_from or not mail_pass:
        result["smtp_test"] = "דולג — חסרים env vars"
        return JSONResponse(result)
    try:
        import asyncio, concurrent.futures
        def _test_smtp():
            with smtplib.SMTP('smtp.office365.com', 587, timeout=15) as s:
                s.ehlo(); s.starttls(); s.ehlo()
                s.login(mail_from, mail_pass)
            return "OK"
        loop = asyncio.get_running_loop()
        with concurrent.futures.ThreadPoolExecutor() as pool:
            smtp_result = await asyncio.wait_for(
                loop.run_in_executor(pool, _test_smtp), timeout=20
            )
        result["smtp_test"] = smtp_result
    except Exception as e:
        result["smtp_test"] = f"שגיאה: {str(e)}"
        print(f"[TEST SMTP ERROR]\n{traceback.format_exc()}")
    return JSONResponse(result)


@app.post("/api/schedule-meeting")
async def schedule_meeting(request: Request):
    """שליחת זימון פגישה (.ics) דרך Brevo REST API"""
    from icalendar import Calendar, Event, vCalAddress, vText
    import uuid, base64, asyncio, concurrent.futures
    import urllib.request, urllib.error
    from datetime import datetime as _dt, timedelta

    user = get_current_user(request)
    if not user or user.get('role') != 'manager':
        return JSONResponse({"error": "אין הרשאה"}, status_code=403)

    data       = await request.json()
    raw_emails = data.get("emails", [])
    if isinstance(raw_emails, str):
        raw_emails = raw_emails.split(',')
    emails = [
        e.strip() for e in raw_emails
        if isinstance(e, str) and '@' in e.strip()
        and '.' in e.strip().split('@')[-1]
        and '\r' not in e and '\n' not in e
    ][:20]

    subject    = (data.get("subject") or f"פגישת {COMPANY_CONFIG.get('company_name', 'SmartConnect')}")[:200]
    notes      = (data.get("notes") or "")[:1000]
    date_str   = data.get("date", "")
    time_str   = data.get("time") or "10:00"
    store_name = (data.get("store_name") or "")[:200]

    try:
        duration = min(480, max(15, int(data.get("duration") or 60)))
    except (ValueError, TypeError):
        duration = 60

    if not emails:
        return JSONResponse({"error": "נא להזין לפחות כתובת מייל אחת תקינה"}, status_code=400)
    if not date_str:
        return JSONResponse({"error": "נא לבחור תאריך"}, status_code=400)

    brevo_key = os.environ.get("BREVO_API_KEY", "")
    mail_from = os.environ.get("MAIL_FROM", COMPANY_CONFIG.get("manager_email", ""))
    if not brevo_key:
        return JSONResponse({"error": "BREVO_API_KEY חסר בהגדרות השרת"}, status_code=500)

    try:
        dt_start = _dt.strptime(f"{date_str} {time_str}", "%Y-%m-%d %H:%M").replace(tzinfo=_IL_TZ)
    except ValueError:
        return JSONResponse({"error": "תאריך או שעה לא תקינים"}, status_code=400)
    dt_end = dt_start + timedelta(minutes=duration)

    cal = Calendar()
    cal.add('prodid', f'-//{COMPANY_CONFIG.get("company_name", "SmartConnect")}//Sales App//HE')
    cal.add('version', '2.0')
    cal.add('method', 'REQUEST')

    event = Event()
    event.add('summary', subject)
    event.add('description', (f"{notes}\n\nחנות: {store_name}".strip() if store_name else notes))
    event.add('dtstart', dt_start)
    event.add('dtend', dt_end)
    event.add('uid', str(uuid.uuid4()))
    organizer = vCalAddress(f"mailto:{mail_from}")
    organizer.params['cn'] = vText(COMPANY_CONFIG.get('company_name', 'SmartConnect'))
    event.add('organizer', organizer, encode=0)
    for email in emails:
        attendee = vCalAddress(f"mailto:{email}")
        attendee.params['cn'] = vText(email)
        attendee.params['ROLE'] = vText('REQ-PARTICIPANT')
        event.add('attendee', attendee, encode=0)
    cal.add_component(event)

    ics_bytes = cal.to_ical()

    body = f"שלום,\n\nהוזמנת לפגישה: {subject}\n"
    if store_name:
        body += f"חנות: {store_name}\n"
    body += f"תאריך: {date_str} {time_str} | משך: {duration} דקות\n"
    if notes:
        body += f"הערות: {notes}\n"
    body += "\nמצורף קובץ זימון לפגישה (.ics) — פתח אותו להוספה ליומן."

    ics_b64 = base64.b64encode(ics_bytes).decode()

    def _send():
        import json as _json
        payload = _json.dumps({
            "sender": {"name": COMPANY_CONFIG.get('company_name', 'SmartConnect'), "email": mail_from},
            "to": [{"email": e} for e in emails],
            "subject": subject,
            "textContent": body,
            "attachment": [{"name": "meeting.ics", "content": ics_b64}]
        }).encode('utf-8')
        req = urllib.request.Request(
            "https://api.brevo.com/v3/smtp/email",
            data=payload,
            headers={"api-key": brevo_key, "Content-Type": "application/json"},
            method="POST"
        )
        with urllib.request.urlopen(req, timeout=20) as resp:
            return resp.read()

    try:
        loop = asyncio.get_running_loop()
        with concurrent.futures.ThreadPoolExecutor() as pool:
            await asyncio.wait_for(loop.run_in_executor(pool, _send), timeout=25)
        return JSONResponse({"status": "ok", "message": f"הזימון נשלח ל-{', '.join(emails)}"})
    except asyncio.TimeoutError:
        return JSONResponse({"error": "פסק זמן בחיבור ל-Brevo"}, status_code=500)
    except Exception as e:
        logging.error("[SCHEDULE MEETING ERROR] %s", e)
        return JSONResponse({"error": "שגיאה בשליחת הזימון — נסה שוב"}, status_code=500)


@app.post("/api/report/email")
async def email_report(request: Request):
    """שליחת דוח אקסל כקובץ מצורף למייל (Office 365 SMTP)"""
    import traceback, smtplib, asyncio, concurrent.futures
    from email.mime.multipart import MIMEMultipart
    from email.mime.base import MIMEBase
    from email.mime.text import MIMEText
    from email import encoders

    user = get_current_user(request)
    if not user or user['role'] != 'manager':
        return JSONResponse({"error": "אין הרשאה"}, status_code=403)

    try:
        data = await request.json()
    except Exception:
        return JSONResponse({"error": "בקשה לא תקינה"}, status_code=400)

    rtype      = data.get("type", "")
    to_email   = data.get("email", "").strip()
    from_date  = data.get("from_date", "")
    to_date    = data.get("to_date", "")
    agent_ids_raw  = data.get("agent_ids", data.get("agent_id"))
    cust_ids_raw   = data.get("customer_ids", data.get("customer_id"))

    if not to_email:
        return JSONResponse({"error": "נא להזין כתובת מייל"}, status_code=400)

    mail_from = os.environ.get("MAIL_FROM", "")
    mail_pass = os.environ.get("MAIL_PASSWORD", "")
    if not mail_from or not mail_pass:
        return JSONResponse({"error": "הגדרות MAIL_FROM / MAIL_PASSWORD חסרות בסביבת השרת"}, status_code=500)

    def _parse_ids(raw):
        if raw is None: return []
        if isinstance(raw, int): return [raw]
        return [int(x.strip()) for x in str(raw).split(',') if x.strip().isdigit()]

    try:
        db = get_db()
        if rtype == "agent":
            ids = _parse_ids(agent_ids_raw)
            if not ids: db.close(); return JSONResponse({"error": "לא נבחרו סוכנים"}, status_code=400)
            headers    = ["תאריך","שם חנות","עיר","אזור","כניסה","יציאה","משך (דק')","סוג","הערות"]
            col_widths = [14,28,14,14,10,10,11,10,30]
            sheets = []
            names = []
            for aid in ids:
                agent = db.execute("SELECT * FROM users WHERE id=?", (aid,)).fetchone()
                if not agent: continue
                rows, total_min = _agent_visit_rows(db, aid, from_date, to_date)
                h, m = total_min // 60, total_min % 60
                sheets.append({'sheet_name': agent['name'], 'title': f"דוח ביקורים - {agent['name']}",
                    'subtitle': f"תקופה: {from_date} עד {to_date}", 'col_widths': col_widths,
                    'headers': headers, 'rows': rows,
                    'summary': [("סה\"כ ביקורים", len(rows)), ("סה\"כ זמן שהייה", f"{h}:{m:02d} שעות")]})
                names.append(agent['name'])
            db.close()
            if not sheets: return JSONResponse({"error": "לא נמצאו סוכנים"}, status_code=404)
            buf = _build_excel_multi(sheets) if len(sheets) > 1 else _build_excel(
                sheets[0]['title'], sheets[0]['subtitle'], col_widths, headers, sheets[0]['rows'], sheets[0]['summary'])
            label = ', '.join(names)
            subject  = f"דוח {COMPANY_CONFIG.get('company_name', 'SmartConnect')} — {label} | {from_date} עד {to_date}"
            filename = f"agents_report_{from_date}_{to_date}.xlsx"

        elif rtype == "customer":
            ids = _parse_ids(cust_ids_raw)
            if not ids: db.close(); return JSONResponse({"error": "לא נבחרו לקוחות"}, status_code=400)
            headers    = ["תאריך","סוכן","כניסה","יציאה","משך (דק')","סוג","הערות"]
            col_widths = [14,18,10,10,11,10,35]
            sheets = []
            names = []
            for cid in ids:
                customer = db.execute("SELECT * FROM customers WHERE id=?", (cid,)).fetchone()
                if not customer: continue
                rows, total_min = _customer_visit_rows(db, cid, from_date, to_date)
                h, m = total_min // 60, total_min % 60
                cname = customer['name']; ccity = customer['city'] or ''
                sheets.append({'sheet_name': f"{cname[:20]}{' '+ccity if ccity else ''}"[:31],
                    'title': f"דוח ביקורים - {cname}",
                    'subtitle': f"תקופה: {from_date} עד {to_date} | {ccity} | {customer['region'] or ''}",
                    'col_widths': col_widths, 'headers': headers, 'rows': rows,
                    'summary': [("סה\"כ ביקורים", len(rows)), ("סה\"כ זמן שהייה", f"{h}:{m:02d} שעות")]})
                names.append(cname)
            db.close()
            if not sheets: return JSONResponse({"error": "לא נמצאו לקוחות"}, status_code=404)
            buf = _build_excel_multi(sheets) if len(sheets) > 1 else _build_excel(
                sheets[0]['title'], sheets[0]['subtitle'], col_widths, headers, sheets[0]['rows'], sheets[0]['summary'])
            label = ', '.join(names)
            subject  = f"דוח {COMPANY_CONFIG.get('company_name', 'SmartConnect')} — {label} | {from_date} עד {to_date}"
            filename = f"customers_report_{from_date}_{to_date}.xlsx"
        else:
            return JSONResponse({"error": "סוג דוח לא תקין"}, status_code=400)

        # Build email
        msg = MIMEMultipart()
        msg['From']    = mail_from
        msg['To']      = to_email
        msg['Subject'] = subject
        body = (f"שלום,\n\nמצורף דוח ביקורים מאפליקציית {COMPANY_CONFIG.get('company_name', 'SmartConnect')}.\n\n"
                f"תקופה: {from_date} עד {to_date}\n\nבברכה,\nמערכת {COMPANY_CONFIG.get('company_name', 'SmartConnect')}")
        msg.attach(MIMEText(body, 'plain', 'utf-8'))

        buf.seek(0)
        part = MIMEBase('application', 'vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        part.set_payload(buf.read())
        encoders.encode_base64(part)
        fname_encoded = quote(filename, safe='')
        part.add_header('Content-Disposition', 'attachment',
                        **{'filename*': f"UTF-8''{fname_encoded}"})
        msg.attach(part)

        # SMTP ב-thread נפרד עם timeout — לא חוסם את event loop
        def _do_send():
            with smtplib.SMTP('smtp.office365.com', 587, timeout=20) as smtp:
                smtp.ehlo()
                smtp.starttls()
                smtp.ehlo()
                smtp.login(mail_from, mail_pass)
                smtp.send_message(msg)

        loop = asyncio.get_running_loop()
        with concurrent.futures.ThreadPoolExecutor() as pool:
            await asyncio.wait_for(
                loop.run_in_executor(pool, _do_send),
                timeout=30
            )

        return JSONResponse({"status": "ok", "message": f"הדוח נשלח בהצלחה ל-{to_email}"})

    except asyncio.TimeoutError:
        print("[EMAIL REPORT] SMTP timeout after 30s")
        return JSONResponse({"error": "פסק זמן בחיבור ל-SMTP — בדוק שהחשבון מאפשר SMTP AUTH ב-Microsoft 365 Admin"}, status_code=500)
    except Exception as e:
        print(f"[EMAIL REPORT ERROR]\n{traceback.format_exc()}")
        err_str = str(e)
        # הסבר ברור לשגיאות נפוצות
        if "535" in err_str or "Authentication" in err_str:
            friendly = "שגיאת אימות — שם משתמש/סיסמה שגויים, או ש-SMTP AUTH לא מופעל בחשבון"
        elif "534" in err_str or "SmtpClientAuthentication" in err_str:
            friendly = "SMTP AUTH מושבת לחשבון זה — יש להפעיל אותו ב-Microsoft 365 Admin Center"
        elif "Connection" in err_str or "timed out" in err_str:
            friendly = "לא ניתן להתחבר לשרת SMTP — חסימת פורט 587 ב-Railway"
        else:
            friendly = err_str
        return JSONResponse({"error": f"שגיאה בשליחת המייל: {friendly}", "raw": err_str}, status_code=500)


@app.get("/api/trigger-eod")
async def trigger_eod(request: Request):
    """הפעלה ידנית של דוח סוף יום (לבדיקה)"""
    user = get_current_user(request)
    if not user or user['role'] != 'manager':
        return JSONResponse({"error": "אין הרשאה"}, status_code=403)
    send_eod_report()
    return JSONResponse({"status": "ok", "message": "דוח נשלח"})


@app.get("/manager/history", response_class=HTMLResponse)
async def manager_history(request: Request):
    user = get_current_user(request)
    if not user or user['role'] != 'manager':
        return RedirectResponse("/")
    db = get_db()
    agents = db.execute("SELECT id, name FROM users WHERE role='agent' ORDER BY name").fetchall()
    db.close()
    return templates.TemplateResponse("history.html", {
        "request": request,
        "user": user,
        "agents": [dict(a) for a in agents],
        "today": now_il().date(),
        "timedelta": timedelta,
        "cfg": COMPANY_CONFIG,
    })


@app.get("/api/history/visits")
async def api_history_visits(
    request: Request,
    agent_id: str = "",
    customer_q: str = "",
    date_from: str = "",
    date_to: str = ""
):
    user = get_current_user(request)
    if not user:
        raise HTTPException(status_code=401)
    if user['role'] != 'manager':
        raise HTTPException(status_code=403)

    db = get_db()
    try:
        conditions = []
        params = []

        if agent_id and agent_id != "all":
            conditions.append("v.user_id = ?")
            params.append(int(agent_id))
        if date_from:
            conditions.append("v.visit_date >= ?")
            params.append(date_from)
        if date_to:
            conditions.append("v.visit_date <= ?")
            params.append(date_to)
        if customer_q:
            conditions.append("c.name LIKE ?")
            params.append(f"%{customer_q}%")

        where = ("WHERE " + " AND ".join(conditions)) if conditions else ""
        sql_history = (
            "SELECT v.id, v.visit_date, v.check_in_time, v.check_out_time,"
            " v.duration_minutes, v.notes, v.customer_id,"
            " u.name as agent_name, c.name as customer_name, c.city, c.region"
            " FROM visits v"
            " JOIN users u ON v.user_id = u.id"
            " JOIN customers c ON v.customer_id = c.id"
            " " + where +
            " ORDER BY v.visit_date DESC, v.check_in_time DESC"
            " LIMIT 200"
        )
        rows = db.execute(sql_history, params).fetchall()

        result = []
        for r in rows:
            result.append({
                "visit_id":      r["id"],
                "date":          r["visit_date"],
                "agent_name":    r["agent_name"],
                "customer_id":   r["customer_id"],
                "customer_name": r["customer_name"],
                "city":          r["city"] or "",
                "region":        r["region"] or "",
                "check_in":      (r["check_in_time"]  or "")[:5],
                "check_out":     (r["check_out_time"] or "")[:5],
                "duration":      r["duration_minutes"] or 0,
                "notes":         r["notes"] or "",
            })
        return JSONResponse(result)
    finally:
        db.close()


@app.get("/manager/daily-summary", response_class=HTMLResponse)
async def daily_summary(request: Request, date: str = None):
    """סיכום יומי לכל עובד — מי ביקר, מי לא, מי לא עדכן יציאה אחרי 17:00"""
    user = get_current_user(request)
    if not user or user['role'] != 'manager':
        return RedirectResponse("/")

    try:
        selected_date = datetime.strptime(date, "%Y-%m-%d").date() if date else now_il().date()
    except ValueError:
        selected_date = now_il().date()

    now_hour     = now_il().hour
    is_after_17  = now_hour >= 17
    hebrew_day   = get_hebrew_day(selected_date)

    db = get_db()
    agents = db.execute("SELECT * FROM users WHERE role='agent' ORDER BY name").fetchall()

    summary = []
    for agent in agents:
        week_num = get_week_for_agent(selected_date, agent)
        week_col = f"week_{week_num}"
        regions = [r.strip() for r in agent['regions'].split(',') if r.strip()] or ['__none__']
        if regions == ['__none__']:
            continue
        placeholders = ','.join(['?' for _ in regions])

        # חנויות מתוכננות להיום
        sql_scheduled = (
            "SELECT * FROM customers"
            " WHERE region IN (" + placeholders + ")"
            " AND assigned_visit_day=?"
            " AND " + week_col + "=1"
            " ORDER BY region, name"
        )
        scheduled = db.execute(
            sql_scheduled,
            regions + [hebrew_day]
        ).fetchall()

        # ביקורים שבוצעו
        visits = db.execute(
            "SELECT * FROM visits WHERE user_id=? AND visit_date=?",
            (agent['id'], str(selected_date))
        ).fetchall()
        visits_map = {v['customer_id']: dict(v) for v in visits}

        done, in_progress, missed, pending = [], [], [], []
        for c in scheduled:
            v = visits_map.get(c['id'])
            if v and v['check_out_time']:
                done.append({'store': dict(c), 'visit': v})
            elif v and v['check_in_time']:
                in_progress.append({'store': dict(c), 'visit': v})
            elif is_after_17 and selected_date == now_il().date():
                missed.append({'store': dict(c), 'visit': None})
            else:
                pending.append({'store': dict(c), 'visit': None})

        summary.append({
            'agent': dict(agent),
            'done':        done,
            'in_progress': in_progress,
            'missed':      missed,
            'pending':     pending,
            'total':       len(scheduled),
        })

    db.close()
    prev_date = (selected_date - timedelta(days=1)).strftime("%Y-%m-%d")
    next_date = (selected_date + timedelta(days=1)).strftime("%Y-%m-%d")

    return templates.TemplateResponse("daily_summary.html", {
        "request": request,
        "user": user,
        "summary": summary,
        "selected_date": selected_date,
        "selected_date_str": str(selected_date),
        "prev_date": prev_date,
        "next_date": next_date,
        "today": now_il().date(),
        "is_after_17": is_after_17,
        "cfg": COMPANY_CONFIG,
    })


@app.get("/manager/agent/{agent_id}", response_class=HTMLResponse)
async def manager_agent_view(request: Request, agent_id: int, date: str = None, status: str = ""):
    user = get_current_user(request)
    if not user or user['role'] != 'manager':
        return RedirectResponse("/")

    try:
        selected_date = datetime.strptime(date, "%Y-%m-%d").date() if date else now_il().date()
    except ValueError:
        selected_date = now_il().date()

    hebrew_day = get_hebrew_day(selected_date)

    db = get_db()
    agent = db.execute("SELECT * FROM users WHERE id=?", (agent_id,)).fetchone()
    if not agent:
        db.close()
        raise HTTPException(404)

    week_num   = get_week_for_agent(selected_date, agent)
    week_col   = f"week_{week_num}"
    regions = [r.strip() for r in agent['regions'].split(',') if r.strip()] or ['__none__']
    placeholders = ','.join(['?' for _ in regions])

    # פילטר לפי לוח החדש (assigned_visit_day + שבוע)
    sql_mgr_agent = (
        "SELECT * FROM customers"
        " WHERE region IN (" + placeholders + ")"
        " AND assigned_visit_day=?"
        " AND " + week_col + "=1"
        " ORDER BY region, name"
    )
    customers = db.execute(
        sql_mgr_agent,
        regions + [hebrew_day]
    ).fetchall()

    # Fallback לשדה הישן אם אין תוצאות
    if not customers:
        sql_mgr_agent_fallback = (
            "SELECT * FROM customers WHERE region IN (" + placeholders + ") AND visit_day=? ORDER BY region, name"
        )
        customers = db.execute(
            sql_mgr_agent_fallback,
            regions + [hebrew_day]
        ).fetchall()

    visits_today = db.execute(
        "SELECT * FROM visits WHERE user_id=? AND visit_date=?",
        (agent_id, str(selected_date))
    ).fetchall()

    eilat_customers = [dict(c) for c in db.execute(
        "SELECT * FROM customers WHERE region='שיראל-אילת' ORDER BY name"
    ).fetchall()] if agent['username'] in ('shirel', 'shirael') else []

    db.close()

    visits_map = {v['customer_id']: dict(v) for v in visits_today}

    # פילטר לפי סטטוס (מגיע מלחיצה על עוגה בדף מנהל)
    status_label = ""
    if status == 'done':
        customers = [c for c in customers if visits_map.get(c['id']) and
                     (visits_map[c['id']].get('check_out_time') or visits_map[c['id']].get('is_phone'))]
        status_label = "✅ הושלמו"
    elif status == 'active':
        customers = [c for c in customers if visits_map.get(c['id']) and
                     visits_map[c['id']].get('check_in_time') and
                     not visits_map[c['id']].get('check_out_time') and
                     not visits_map[c['id']].get('is_phone')]
        status_label = "🟡 כרגע בפנים"
    elif status == 'red':
        customers = [c for c in customers if not visits_map.get(c['id']) or
                     not visits_map[c['id']].get('check_in_time')]
        status_label = "🔴 לא ביקרו"

    prev_date = (selected_date - timedelta(days=1)).strftime("%Y-%m-%d")
    next_date = (selected_date + timedelta(days=1)).strftime("%Y-%m-%d")

    # לוח חודשי
    import calendar
    year, month = selected_date.year, selected_date.month
    first_day = selected_date.replace(day=1)
    first_weekday = (first_day.weekday() + 1) % 7
    days_in_month = calendar.monthrange(year, month)[1]
    month_days = [None] * first_weekday
    for d in range(1, days_in_month + 1):
        month_days.append(selected_date.replace(day=d))
    while len(month_days) % 7 != 0:
        month_days.append(None)

    month_name_map = {1:'ינואר',2:'פברואר',3:'מרץ',4:'אפריל',5:'מאי',6:'יוני',
                      7:'יולי',8:'אוגוסט',9:'ספטמבר',10:'אוקטובר',11:'נובמבר',12:'דצמבר'}

    return templates.TemplateResponse("agent.html", {
        "request": request,
        "user": user,
        "viewed_agent": dict(agent),
        "customers": [dict(c) for c in customers],
        "visits_map": visits_map,
        "selected_date": selected_date,
        "selected_date_str": str(selected_date),
        "hebrew_day": hebrew_day,
        "hebrew_day_name": HEBREW_DAY_NAMES.get(hebrew_day, ''),
        "week_num": week_num,
        "prev_date": prev_date,
        "next_date": next_date,
        "week_days": [],
        "month_days": month_days,
        "month_label": f"{month_name_map.get(month,'')} {year}",
        "timedelta": timedelta,
        "today": now_il().date(),
        "is_after_17": now_il().hour >= 17,
        "HEBREW_DAYS": HEBREW_DAYS,
        "TRAFFIC_LIGHT_COLORS": TRAFFIC_LIGHT_COLORS,
        "manager_view": True,
        "eilat_customers": eilat_customers,
        "status_filter": status,
        "status_label": status_label,
        "cfg": COMPANY_CONFIG,
    })

# ─── V2 ROUTES (new UI — originals untouched) ────────────────────────────────

@app.get("/v2/dashboard", response_class=HTMLResponse)
async def agent_dashboard_v2(request: Request, date: str = None):
    user = get_current_user(request)
    if not user:
        return RedirectResponse("/")
    if user['role'] == 'manager':
        return RedirectResponse("/v2/manager")

    try:
        selected_date = datetime.strptime(date, "%Y-%m-%d").date() if date else now_il().date()
    except ValueError:
        selected_date = now_il().date()

    hebrew_day = get_hebrew_day(selected_date)
    week_num   = get_week_for_agent(selected_date, user)
    week_col   = f"week_{week_num}"
    regions = [r.strip() for r in user['regions'].split(',') if r.strip()] or ['__none__']

    db = get_db()
    placeholders = ','.join(['?' for _ in regions])

    sql_customers_v2 = (
        "SELECT * FROM customers"
        " WHERE region IN (" + placeholders + ")"
        " AND assigned_visit_day=?"
        " AND " + week_col + "=1"
        " ORDER BY region, name"
    )
    customers = db.execute(
        sql_customers_v2,
        regions + [hebrew_day]
    ).fetchall()

    if not customers:
        sql_customers_v2_fallback = (
            "SELECT * FROM customers WHERE region IN (" + placeholders + ") AND visit_day=? ORDER BY region, name"
        )
        customers = db.execute(
            sql_customers_v2_fallback,
            regions + [hebrew_day]
        ).fetchall()

    visits_today = db.execute(
        "SELECT * FROM visits WHERE user_id=? AND visit_date=?",
        (user['id'], str(selected_date))
    ).fetchall()
    db.close()

    visits_map = {v['customer_id']: dict(v) for v in visits_today}
    prev_date = (selected_date - timedelta(days=1)).strftime("%Y-%m-%d")
    next_date = (selected_date + timedelta(days=1)).strftime("%Y-%m-%d")

    import calendar
    year, month = selected_date.year, selected_date.month
    first_day = selected_date.replace(day=1)
    first_weekday = (first_day.weekday() + 1) % 7
    days_in_month = calendar.monthrange(year, month)[1]
    month_days = [None] * first_weekday
    for d in range(1, days_in_month + 1):
        month_days.append(selected_date.replace(day=d))
    while len(month_days) % 7 != 0:
        month_days.append(None)

    month_name_map = {1:'ינואר',2:'פברואר',3:'מרץ',4:'אפריל',5:'מאי',6:'יוני',
                      7:'יולי',8:'אוגוסט',9:'ספטמבר',10:'אוקטובר',11:'נובמבר',12:'דצמבר'}

    return templates.TemplateResponse("agent_v2.html", {
        "request": request,
        "user": user,
        "customers": [dict(c) for c in customers],
        "visits_map": visits_map,
        "selected_date": selected_date,
        "selected_date_str": str(selected_date),
        "hebrew_day": hebrew_day,
        "hebrew_day_name": HEBREW_DAY_NAMES.get(hebrew_day, ''),
        "week_num": week_num,
        "prev_date": prev_date,
        "next_date": next_date,
        "week_days": [],
        "month_days": month_days,
        "month_label": f"{month_name_map.get(month,'')} {year}",
        "timedelta": timedelta,
        "today": now_il().date(),
        "is_after_17": now_il().hour >= 17,
        "HEBREW_DAYS": HEBREW_DAYS,
        "TRAFFIC_LIGHT_COLORS": TRAFFIC_LIGHT_COLORS,
        "cfg": COMPANY_CONFIG,
    })


@app.get("/v2/store/{customer_id}", response_class=HTMLResponse)
async def store_detail_v2(request: Request, customer_id: int, date: str = None, agent_id: int = None):
    user = get_current_user(request)
    if not user:
        return RedirectResponse("/")

    try:
        visit_date = datetime.strptime(date, "%Y-%m-%d").date() if date else now_il().date()
    except ValueError:
        visit_date = now_il().date()

    is_manager = (user['role'] == 'manager')

    db = get_db()
    customer = db.execute("SELECT * FROM customers WHERE id=?", (customer_id,)).fetchone()
    if not customer:
        db.close()
        raise HTTPException(404)

    if is_manager and agent_id:
        visit = db.execute(
            "SELECT * FROM visits WHERE user_id=? AND customer_id=? AND visit_date=?",
            (agent_id, customer_id, str(visit_date))
        ).fetchone()
        back_url = f"/v2/manager/agent/{agent_id}?date={visit_date}"
    else:
        visit = db.execute(
            "SELECT * FROM visits WHERE user_id=? AND customer_id=? AND visit_date=?",
            (user['id'], customer_id, str(visit_date))
        ).fetchone()
        back_url = f"/v2/dashboard?date={visit_date}"

    history = db.execute(
        """SELECT v.*, u.name as agent_name FROM visits v
           JOIN users u ON v.user_id = u.id
           WHERE v.customer_id=? ORDER BY v.visit_date DESC, v.check_in_time DESC LIMIT 20""",
        (customer_id,)
    ).fetchall()
    db.close()

    return templates.TemplateResponse("store_v2.html", {
        "request": request,
        "user": user,
        "customer": dict(customer),
        "visit": dict(visit) if visit else None,
        "history": [dict(h) for h in history],
        "visit_date": str(visit_date),
        "is_manager": is_manager,
        "back_url": back_url,
        "HEBREW_DAY_NAMES": HEBREW_DAY_NAMES,
        "cfg": COMPANY_CONFIG,
    })


@app.get("/v2/manager", response_class=HTMLResponse)
async def manager_dashboard_v2(request: Request, date: str = None, agent_id: int = None):
    user = get_current_user(request)
    if not user or user['role'] != 'manager':
        return RedirectResponse("/")

    try:
        selected_date = datetime.strptime(date, "%Y-%m-%d").date() if date else now_il().date()
    except ValueError:
        selected_date = now_il().date()

    db = get_db()
    agents = db.execute("SELECT * FROM users WHERE role='agent' ORDER BY name").fetchall()

    visits_query = "SELECT v.*, u.name as agent_name, c.name as store_name, c.city, c.region FROM visits v JOIN users u ON v.user_id=u.id JOIN customers c ON v.customer_id=c.id WHERE v.visit_date=?"
    params = [str(selected_date)]
    if agent_id:
        visits_query += " AND v.user_id=?"
        params.append(agent_id)
    visits_query += " ORDER BY v.check_in_time DESC"

    visits = db.execute(visits_query, params).fetchall()

    notifications = db.execute(
        "SELECT * FROM notifications ORDER BY created_at DESC LIMIT 50"
    ).fetchall()

    eod_reports = db.execute(
        """SELECT * FROM notifications
           WHERE action='eod_report' AND created_at LIKE ?
           ORDER BY created_at DESC""",
        (str(selected_date) + '%',)
    ).fetchall()

    unread = db.execute("SELECT COUNT(*) FROM notifications WHERE is_read=0").fetchone()[0]

    agent_summary = []
    for agent in agents:
        today_visits = db.execute(
            "SELECT COUNT(*) FROM visits WHERE user_id=? AND visit_date=? AND (check_out_time IS NOT NULL OR is_phone=1)",
            (agent['id'], str(selected_date))
        ).fetchone()[0]
        current = db.execute(
            """SELECT v.*, c.name as store_name FROM visits v JOIN customers c ON v.customer_id=c.id
               WHERE v.user_id=? AND v.visit_date=? AND v.check_in_time IS NOT NULL AND v.check_out_time IS NULL""",
            (agent['id'], str(selected_date))
        ).fetchone()
        last_act = db.execute(
            """SELECT v.check_in_time, v.check_out_time, c.name as store_name
               FROM visits v JOIN customers c ON v.customer_id=c.id
               WHERE v.user_id=? AND v.visit_date=? AND v.check_in_time IS NOT NULL
               ORDER BY COALESCE(v.check_out_time, v.check_in_time) DESC LIMIT 1""",
            (agent['id'], str(selected_date))
        ).fetchone()
        last_activity_time  = None
        last_activity_store = None
        last_activity_type  = None
        if last_act:
            t = last_act['check_out_time'] or last_act['check_in_time']
            last_activity_time  = t[:5] if t else None
            last_activity_store = last_act['store_name']
            last_activity_type  = 'יציאה' if last_act['check_out_time'] else 'כניסה'
        agent_regions = [r.strip() for r in (agent['regions'] or '').split(',') if r.strip()]
        if agent_regions:
            hebrew_day_sel = get_hebrew_day(selected_date)
            week_num_sel   = get_week_for_agent(selected_date, agent)
            week_col_sel   = f"week_{week_num_sel}"
            placeholders   = ','.join(['?' for _ in agent_regions])
            total_customers = db.execute(
                f"SELECT COUNT(*) FROM customers WHERE region IN ({placeholders}) AND assigned_visit_day=? AND {week_col_sel}=1",
                agent_regions + [hebrew_day_sel]
            ).fetchone()[0]
            if total_customers == 0:
                total_customers = db.execute(
                    f"SELECT COUNT(*) FROM customers WHERE region IN ({placeholders}) AND visit_day=?",
                    agent_regions + [hebrew_day_sel]
                ).fetchone()[0]
        else:
            total_customers = 0
        agent_summary.append({
            "id": agent['id'],
            "name": agent['name'],
            "username": agent['username'],
            "regions": agent['regions'],
            "today_visits": today_visits,
            "total_customers": total_customers,
            "current_store": dict(current) if current else None,
            "last_activity_time":  last_activity_time,
            "last_activity_store": last_activity_store,
            "last_activity_type":  last_activity_type,
        })

    db.close()

    prev_date = (selected_date - timedelta(days=1)).strftime("%Y-%m-%d")
    next_date = (selected_date + timedelta(days=1)).strftime("%Y-%m-%d")

    return templates.TemplateResponse("manager_v2.html", {
        "request": request,
        "user": user,
        "agents": [dict(a) for a in agents],
        "agent_summary": agent_summary,
        "eod_reports": [dict(r) for r in eod_reports],
        "visits": [dict(v) for v in visits],
        "notifications": [dict(n) for n in notifications],
        "unread": unread,
        "selected_date": selected_date,
        "selected_date_str": str(selected_date),
        "prev_date": prev_date,
        "next_date": next_date,
        "selected_agent_id": agent_id,
        "today": now_il().date(),
        "manager_email": get_setting("manager_email", COMPANY_CONFIG.get("manager_email", "ranaz@matrix.co.il")),
        "cfg": COMPANY_CONFIG,
    })


@app.get("/v2/manager/agent/{agent_id}", response_class=HTMLResponse)
async def manager_agent_view_v2(request: Request, agent_id: int, date: str = None):
    user = get_current_user(request)
    if not user or user['role'] != 'manager':
        return RedirectResponse("/")

    try:
        selected_date = datetime.strptime(date, "%Y-%m-%d").date() if date else now_il().date()
    except ValueError:
        selected_date = now_il().date()

    hebrew_day = get_hebrew_day(selected_date)

    db = get_db()
    agent = db.execute("SELECT * FROM users WHERE id=?", (agent_id,)).fetchone()
    if not agent:
        db.close()
        raise HTTPException(404)

    week_num   = get_week_for_agent(selected_date, agent)
    week_col   = f"week_{week_num}"
    regions = [r.strip() for r in agent['regions'].split(',') if r.strip()] or ['__none__']
    placeholders = ','.join(['?' for _ in regions])

    sql_mgr_agent_v2 = (
        "SELECT * FROM customers"
        " WHERE region IN (" + placeholders + ")"
        " AND assigned_visit_day=?"
        " AND " + week_col + "=1"
        " ORDER BY region, name"
    )
    customers = db.execute(
        sql_mgr_agent_v2,
        regions + [hebrew_day]
    ).fetchall()

    if not customers:
        sql_mgr_agent_v2_fallback = (
            "SELECT * FROM customers WHERE region IN (" + placeholders + ") AND visit_day=? ORDER BY region, name"
        )
        customers = db.execute(
            sql_mgr_agent_v2_fallback,
            regions + [hebrew_day]
        ).fetchall()

    visits_today = db.execute(
        "SELECT * FROM visits WHERE user_id=? AND visit_date=?",
        (agent_id, str(selected_date))
    ).fetchall()
    db.close()

    visits_map = {v['customer_id']: dict(v) for v in visits_today}
    prev_date = (selected_date - timedelta(days=1)).strftime("%Y-%m-%d")
    next_date = (selected_date + timedelta(days=1)).strftime("%Y-%m-%d")

    import calendar
    year, month = selected_date.year, selected_date.month
    first_day = selected_date.replace(day=1)
    first_weekday = (first_day.weekday() + 1) % 7
    days_in_month = calendar.monthrange(year, month)[1]
    month_days = [None] * first_weekday
    for d in range(1, days_in_month + 1):
        month_days.append(selected_date.replace(day=d))
    while len(month_days) % 7 != 0:
        month_days.append(None)

    month_name_map = {1:'ינואר',2:'פברואר',3:'מרץ',4:'אפריל',5:'מאי',6:'יוני',
                      7:'יולי',8:'אוגוסט',9:'ספטמבר',10:'אוקטובר',11:'נובמבר',12:'דצמבר'}

    return templates.TemplateResponse("agent_v2.html", {
        "request": request,
        "user": user,
        "viewed_agent": dict(agent),
        "customers": [dict(c) for c in customers],
        "visits_map": visits_map,
        "selected_date": selected_date,
        "selected_date_str": str(selected_date),
        "hebrew_day": hebrew_day,
        "hebrew_day_name": HEBREW_DAY_NAMES.get(hebrew_day, ''),
        "week_num": week_num,
        "prev_date": prev_date,
        "next_date": next_date,
        "week_days": [],
        "month_days": month_days,
        "month_label": f"{month_name_map.get(month,'')} {year}",
        "timedelta": timedelta,
        "today": now_il().date(),
        "is_after_17": now_il().hour >= 17,
        "HEBREW_DAYS": HEBREW_DAYS,
        "TRAFFIC_LIGHT_COLORS": TRAFFIC_LIGHT_COLORS,
        "manager_view": True,
        "cfg": COMPANY_CONFIG,
    })
