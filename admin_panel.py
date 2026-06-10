"""
Admin Panel — ניהול לקוחות
/admin/* — נפרד לחלוטין מהאפליקציה הראשית
"""
import json
import logging
import os
from datetime import datetime, timedelta
from typing import Optional

from fastapi import APIRouter, Request, Form, File, UploadFile
from fastapi.responses import HTMLResponse, RedirectResponse, JSONResponse
import io
from fastapi.templating import Jinja2Templates
from itsdangerous import TimestampSigner, BadSignature, SignatureExpired

from database import get_db
from pathlib import Path as _Path

_cfg_path = _Path(__file__).parent / "company_config.json"
COMPANY_CONFIG = json.loads(_cfg_path.read_text(encoding="utf-8")) if _cfg_path.exists() else {}

# ── Constants ────────────────────────────────────────────────────────────────
ADMIN_USER   = os.environ.get("ADMIN_USER",  "Admin")
ADMIN_PASS   = os.environ.get("ADMIN_PASS",  "Admin2026")
_secret      = os.environ.get("SECRET_KEY",  "dev-fallback-key")
_signer      = TimestampSigner(_secret)
ADMIN_COOKIE = "admin_session"

_tpl = Jinja2Templates(directory="templates")
admin_router = APIRouter(prefix="/admin", tags=["admin"])
log = logging.getLogger("admin_panel")

HEB_DAYS = {
    'א': 'ראשון', 'ב': 'שני', 'ג': 'שלישי',
    'ד': 'רביעי', 'ה': 'חמישי'
}

# ── Auth ─────────────────────────────────────────────────────────────────────
def _get_admin(request: Request) -> bool:
    cookie = request.cookies.get(ADMIN_COOKIE)
    if not cookie:
        return False
    try:
        _signer.unsign(cookie, max_age=86400)
        return True
    except (BadSignature, SignatureExpired):
        return False


# ── Cleanup — exported to main.py scheduler ──────────────────────────────────
def cleanup_deleted_customers():
    """מוחק לצמיתות לקוחות שנמחקו לפני יותר מ-30 יום (רץ אוטומטית ב-03:00)"""
    try:
        db = get_db()
        cutoff = (datetime.now() - timedelta(days=30)).isoformat()
        cur = db.execute(
            "DELETE FROM customers WHERE deleted_at IS NOT NULL AND deleted_at != '' AND deleted_at < ?",
            (cutoff,)
        )
        db.commit()
        count = getattr(cur, 'rowcount', 0) or 0
        db.close()
        log.info(f"[admin cleanup] נמחקו {count} לקוחות שפג תוקף שחזורם")
    except Exception as exc:
        log.error(f"[admin cleanup] שגיאה: {exc}")


# ── Helper ───────────────────────────────────────────────────────────────────
def _row_to_weeks(r) -> dict:
    """מחלץ week_1..week_7 מ-Row, עם fallback ל-0 אם העמודה חסרה"""
    result = {}
    for i in range(1, 8):
        try:
            result[f"week_{i}"] = int(r[f"week_{i}"] or 0)
        except Exception:
            result[f"week_{i}"] = 0
    return result


# ── Pages ─────────────────────────────────────────────────────────────────────
@admin_router.get("/", include_in_schema=False)
async def admin_root(request: Request):
    if _get_admin(request):
        return RedirectResponse("/admin/customers")
    return RedirectResponse("/admin/login")


@admin_router.get("/login", response_class=HTMLResponse)
async def admin_login_get(request: Request):
    if _get_admin(request):
        return RedirectResponse("/admin/customers")
    return _tpl.TemplateResponse("admin_login.html", {"request": request, "error": None, "cfg": COMPANY_CONFIG})


@admin_router.post("/login")
async def admin_login_post(
    request: Request,
    username: str = Form(...),
    password: str = Form(...),
):
    if username == ADMIN_USER and password == ADMIN_PASS:
        token = _signer.sign(b"admin").decode()
        resp = RedirectResponse("/admin/customers", status_code=303)
        resp.set_cookie(ADMIN_COOKIE, token, max_age=86400, httponly=True, samesite="lax")
        return resp
    return _tpl.TemplateResponse(
        "admin_login.html",
        {"request": request, "error": "שם משתמש או סיסמה שגויים", "cfg": COMPANY_CONFIG},
        status_code=401,
    )


@admin_router.get("/logout")
async def admin_logout():
    resp = RedirectResponse("/admin/login")
    resp.delete_cookie(ADMIN_COOKIE)
    return resp


@admin_router.get("/customers", response_class=HTMLResponse)
async def admin_customers_page(request: Request):
    if not _get_admin(request):
        return RedirectResponse("/admin/login")
    return _tpl.TemplateResponse("admin_customers.html", {"request": request, "cfg": COMPANY_CONFIG})


@admin_router.get("/trash", response_class=HTMLResponse)
async def admin_trash_page(request: Request):
    if not _get_admin(request):
        return RedirectResponse("/admin/login")
    return _tpl.TemplateResponse("admin_trash.html", {"request": request, "cfg": COMPANY_CONFIG})


# ── API — Regions ─────────────────────────────────────────────────────────────
@admin_router.get("/api/regions")
async def api_regions(request: Request):
    if not _get_admin(request):
        return JSONResponse({"error": "אין הרשאה"}, status_code=401)
    db = get_db()
    rows = db.execute(
        "SELECT DISTINCT region FROM customers "
        "WHERE region IS NOT NULL AND region != '' "
        "ORDER BY region"
    ).fetchall()
    db.close()
    return JSONResponse([r["region"] for r in rows])


# ── API — Customers List ───────────────────────────────────────────────────────
@admin_router.get("/api/customers")
async def api_customers_list(
    request: Request,
    q: str = "",
    agent: str = "",
    day: str = "",
):
    if not _get_admin(request):
        return JSONResponse({"error": "אין הרשאה"}, status_code=401)

    db = get_db()
    sql = (
        "SELECT id, card_code, name, city, address, region, "
        "assigned_visit_day, week_1, week_2, week_3, week_4, "
        "week_5, week_6, week_7, delivery_day "
        "FROM customers "
        "WHERE (deleted_at IS NULL OR deleted_at = '')"
    )
    params: list = []

    if q:
        sql += " AND (name LIKE ? OR city LIKE ? OR card_code LIKE ?)"
        like = f"%{q}%"
        params += [like, like, like]
    if agent:
        sql += " AND region LIKE ?"
        params.append(f"%{agent}%")
    if day:
        sql += " AND assigned_visit_day = ?"
        params.append(day)

    sql += " ORDER BY region, assigned_visit_day, name LIMIT 500"

    rows = db.execute(sql, params).fetchall()
    db.close()

    result = []
    for r in rows:
        w = _row_to_weeks(r)
        weeks_active = [i for i in range(1, 8) if w[f"week_{i}"]]
        result.append({
            "id":            r["id"],
            "card_code":     r["card_code"]         or "",
            "name":          r["name"],
            "city":          r["city"]              or "",
            "address":       r["address"]           or "",
            "region":        r["region"]            or "",
            "day":           r["assigned_visit_day"] or "",
            "day_name":      HEB_DAYS.get(r["assigned_visit_day"] or "", "—"),
            "week_1":        w["week_1"],
            "week_2":        w["week_2"],
            "week_3":        w["week_3"],
            "week_4":        w["week_4"],
            "week_5":        w["week_5"],
            "week_6":        w["week_6"],
            "week_7":        w["week_7"],
            "weeks_display": " ".join([f"ש{i}" for i in weeks_active]) if weeks_active else "—",
            "delivery_day":  r["delivery_day"]      or "",
        })
    return JSONResponse(result)


# ── API — Create Customer ─────────────────────────────────────────────────────
@admin_router.post("/api/customers")
async def api_customers_create(request: Request):
    if not _get_admin(request):
        return JSONResponse({"error": "אין הרשאה"}, status_code=401)

    data         = await request.json()
    card_code    = (data.get("card_code") or "").strip()
    name         = (data.get("name")      or "").strip()
    city         = (data.get("city")      or "").strip()
    address      = (data.get("address")   or "").strip()
    region       = (data.get("region")    or "").strip()
    assigned_day = (data.get("day")       or "").strip() or None
    delivery_day = (data.get("delivery_day") or "").strip() or None
    weeks        = data.get("weeks", [])   # list of ints e.g. [1, 3, 5]

    if not name:
        return JSONResponse({"error": "שם לקוח הוא שדה חובה"}, status_code=400)

    w = {f"week_{i}": (1 if i in weeks else 0) for i in range(1, 8)}

    db = get_db()
    try:
        cur = db.execute(
            """INSERT INTO customers
               (card_code, name, city, address, region, assigned_visit_day,
                week_1, week_2, week_3, week_4, week_5, week_6, week_7,
                delivery_day, visit_day, traffic_light)
               VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
            (card_code, name, city, address, region, assigned_day,
             w["week_1"], w["week_2"], w["week_3"], w["week_4"],
             w["week_5"], w["week_6"], w["week_7"],
             delivery_day, assigned_day, "ירוק"),
        )
        new_id = cur.lastrowid
        db.commit()
    finally:
        db.close()

    return JSONResponse({"ok": True, "id": new_id})


# ── API — Update Customer ─────────────────────────────────────────────────────
@admin_router.put("/api/customers/{cid}")
async def api_customers_update(request: Request, cid: int):
    if not _get_admin(request):
        return JSONResponse({"error": "אין הרשאה"}, status_code=401)

    data         = await request.json()
    card_code    = (data.get("card_code")    or "").strip()
    name         = (data.get("name")         or "").strip()
    city         = (data.get("city")         or "").strip()
    address      = (data.get("address")      or "").strip()
    region       = (data.get("region")       or "").strip()
    assigned_day = (data.get("day")          or "").strip() or None
    delivery_day = (data.get("delivery_day") or "").strip() or None
    weeks        = data.get("weeks", [])

    if not name:
        return JSONResponse({"error": "שם לקוח הוא שדה חובה"}, status_code=400)

    w = {f"week_{i}": (1 if i in weeks else 0) for i in range(1, 8)}

    db = get_db()
    try:
        db.execute(
            """UPDATE customers SET
               card_code=?, name=?, city=?, address=?, region=?,
               assigned_visit_day=?,
               week_1=?, week_2=?, week_3=?, week_4=?,
               week_5=?, week_6=?, week_7=?,
               delivery_day=?
               WHERE id=?""",
            (card_code, name, city, address, region, assigned_day,
             w["week_1"], w["week_2"], w["week_3"], w["week_4"],
             w["week_5"], w["week_6"], w["week_7"],
             delivery_day, cid),
        )
        db.commit()
    finally:
        db.close()

    return JSONResponse({"ok": True})


# ── API — Soft Delete ─────────────────────────────────────────────────────────
@admin_router.delete("/api/customers/{cid}")
async def api_customers_delete(request: Request, cid: int):
    if not _get_admin(request):
        return JSONResponse({"error": "אין הרשאה"}, status_code=401)

    db = get_db()
    try:
        row = db.execute(
            "SELECT name, assigned_visit_day, week_1, week_2, week_3, week_4, "
            "week_5, week_6, week_7, delivery_day FROM customers WHERE id=?",
            (cid,),
        ).fetchone()

        if not row:
            return JSONResponse({"error": "לקוח לא נמצא"}, status_code=404)

        w = _row_to_weeks(row)
        backup = json.dumps({
            "assigned_visit_day": row["assigned_visit_day"],
            "delivery_day":       row["delivery_day"],
            **w,
        }, ensure_ascii=False)

        db.execute(
            """UPDATE customers SET
               assigned_visit_day=NULL,
               week_1=0, week_2=0, week_3=0, week_4=0,
               week_5=0, week_6=0, week_7=0,
               deleted_at=?, deleted_backup=?
               WHERE id=?""",
            (datetime.now().isoformat(), backup, cid),
        )
        db.commit()
    finally:
        db.close()

    return JSONResponse({"ok": True})


# ── API — Restore ─────────────────────────────────────────────────────────────
@admin_router.post("/api/customers/{cid}/restore")
async def api_customers_restore(request: Request, cid: int):
    if not _get_admin(request):
        return JSONResponse({"error": "אין הרשאה"}, status_code=401)

    db = get_db()
    try:
        row = db.execute(
            "SELECT deleted_backup FROM customers "
            "WHERE id=? AND deleted_at IS NOT NULL AND deleted_at != ''",
            (cid,),
        ).fetchone()

        if not row or not row["deleted_backup"]:
            return JSONResponse({"error": "לא נמצא לקוח מחוק"}, status_code=404)

        b = json.loads(row["deleted_backup"])
        db.execute(
            """UPDATE customers SET
               assigned_visit_day=?,
               week_1=?, week_2=?, week_3=?, week_4=?,
               week_5=?, week_6=?, week_7=?,
               delivery_day=?,
               deleted_at=NULL, deleted_backup=NULL
               WHERE id=?""",
            (b.get("assigned_visit_day"),
             b.get("week_1", 0), b.get("week_2", 0), b.get("week_3", 0), b.get("week_4", 0),
             b.get("week_5", 0), b.get("week_6", 0), b.get("week_7", 0),
             b.get("delivery_day"), cid),
        )
        db.commit()
    finally:
        db.close()

    return JSONResponse({"ok": True})


# ── API — Trash List ──────────────────────────────────────────────────────────
@admin_router.post("/api/import-excel")
async def api_import_excel(request: Request, file: UploadFile = File(...)):
    """ייבוא לקוחות מ-Excel — POST /admin/api/import-excel"""
    if not _get_admin(request):
        return JSONResponse({"error": "אין הרשאה"}, status_code=401)

    try:
        import openpyxl
    except ImportError:
        return JSONResponse({"error": "חסר מודול openpyxl — הרץ: pip install openpyxl"}, status_code=500)

    contents = await file.read()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(contents), data_only=True)
    except Exception as e:
        return JSONResponse({"error": f"שגיאה בקריאת קובץ Excel: {e}"}, status_code=400)

    ws = wb.active
    headers_row = [str(c.value or "").strip().lower() for c in ws[1]]

    # מיפוי שמות עמודות ← גמיש
    COL_MAP = {
        "card_code":        ["card_code", "קוד לקוח", "קוד", "cardcode", "bp_code"],
        "name":             ["name", "שם", "שם לקוח", "שם חנות", "customer_name"],
        "city":             ["city", "עיר", "city"],
        "address":          ["address", "כתובת", "street"],
        "region":           ["region", "אזור", "סוכן"],
        "delivery_day":     ["delivery_day", "יום אספקה", "delivery"],
        "x_days":           ["x_days", "תדירות", "x days", "frequency"],
        "visit_day":        ["visit_day", "יום ביקור", "visit day"],
        "traffic_light":    ["traffic_light", "רמזור", "traffic"],
        "assigned_visit_day": ["assigned_visit_day", "יום מוקצה", "assigned day"],
    }

    def find_col(field):
        aliases = COL_MAP.get(field, [])
        for alias in aliases:
            if alias.lower() in headers_row:
                return headers_row.index(alias.lower())
        return None

    col_idx = {field: find_col(field) for field in COL_MAP}

    if col_idx["name"] is None:
        return JSONResponse({"error": "לא נמצאה עמודת שם לקוח — ודא שיש עמודה 'name' או 'שם לקוח'"}, status_code=400)

    def cell(row, field):
        idx = col_idx.get(field)
        if idx is None:
            return None
        v = row[idx].value
        return str(v).strip() if v is not None else None

    db = get_db()
    inserted = 0
    skipped  = 0
    errors   = []

    for row_num, row in enumerate(ws.iter_rows(min_row=2), start=2):
        name = cell(row, "name")
        if not name:
            skipped += 1
            continue
        try:
            db.execute(
                """INSERT INTO customers
                   (card_code, name, city, address, region, delivery_day,
                    x_days, visit_day, traffic_light, assigned_visit_day)
                   VALUES (?,?,?,?,?,?,?,?,?,?)""",
                (
                    cell(row, "card_code"),
                    name,
                    cell(row, "city"),
                    cell(row, "address"),
                    cell(row, "region"),
                    cell(row, "delivery_day"),
                    int(cell(row, "x_days") or 0) if cell(row, "x_days") else None,
                    cell(row, "visit_day"),
                    cell(row, "traffic_light"),
                    cell(row, "assigned_visit_day"),
                )
            )
            inserted += 1
        except Exception as e:
            errors.append(f"שורה {row_num}: {e}")

    db.commit()
    db.close()

    return JSONResponse({
        "ok": True,
        "inserted": inserted,
        "skipped": skipped,
        "errors": errors[:10],  # מקסימום 10 שגיאות בתצוגה
        "message": f"יובאו {inserted} לקוחות בהצלחה"
    })


@admin_router.get("/api/trash")
async def api_trash_list(request: Request):
    if not _get_admin(request):
        return JSONResponse({"error": "אין הרשאה"}, status_code=401)

    db = get_db()
    rows = db.execute(
        "SELECT id, card_code, name, city, region, deleted_at, deleted_backup "
        "FROM customers "
        "WHERE deleted_at IS NOT NULL AND deleted_at != '' "
        "ORDER BY deleted_at DESC"
    ).fetchall()
    db.close()

    now    = datetime.now()
    result = []
    for r in rows:
        try:
            del_dt         = datetime.fromisoformat(r["deleted_at"])
            days_elapsed   = (now - del_dt).days
            days_remaining = max(0, 30 - days_elapsed)
            del_date_str   = del_dt.strftime("%d/%m/%Y")
        except Exception:
            days_remaining = 30
            del_date_str   = (r["deleted_at"] or "")[:10]

        backup   = {}
        if r["deleted_backup"]:
            try:
                backup = json.loads(r["deleted_backup"])
            except Exception:
                pass

        result.append({
            "id":            r["id"],
            "card_code":     r["card_code"]  or "",
            "name":          r["name"],
            "city":          r["city"]       or "",
            "region":        r["region"]     or "",
            "deleted_date":  del_date_str,
            "days_remaining": days_remaining,
            "original_day":  HEB_DAYS.get(backup.get("assigned_visit_day", ""), "—"),
        })

    return JSONResponse(result)
